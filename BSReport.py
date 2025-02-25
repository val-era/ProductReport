import base64
import io

import pandas as pd
import openpyxl
from openpyxl.cell.cell import Cell
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from UploadPhoto import GetImage
from PIL import JpegImagePlugin

class Processing:

    def __init__(self):
        self.file_name = "TestFile.xlsx"
        self.category_df = None

    def process(self, file_name):
        self.file_name = file_name
        start_cr_df = CreateStoreDF()  # Класс чтения файла и разбивки на DF магазина
        self.category_df = start_cr_df.create_df(self.file_name)
        self.excel_file()

    def excel_file(self):
        JpegImagePlugin._getmp = lambda: None
        wb = openpyxl.Workbook()
        worksheet = wb.active
        wb.remove(worksheet)
        for store in self.category_df:
            # Создание листов для записи в файл
            wb.create_sheet(title=store)
            category_df = self.category_df[store]
            ws = wb[store]
            ws.alignment = Alignment(wrapText=True)
            # Составляем общую
            ttl_df = category_df["TTL"]
            ttl_list = ttl_df.values.tolist()
            header = ["TOTAL"]

            def styled_cells(header):
                for c in header:
                    c = Cell(ws, column="A", row=1, value=c)
                    c.font = Font(bold=True)
                    c.fill = PatternFill(patternType='solid',
                                         fgColor='e6e6e6')
                    yield c

            ws.append(styled_cells(header))
            headers = ['Артикул', "IMG", 'Модель ANTA', 'Сезон', 'RRP', 'Пол',
                       'Коллекция', 'Кол-во\nпродаж', 'Сумма со скидкой', "Moscow Category\nNet Sales rank"]

            def styled_cells(headers):
                for c in headers:
                    c = Cell(ws, column="A", row=1, value=c)
                    c.font = Font(bold=True)
                    c.fill = PatternFill(patternType='solid',
                                         fgColor='e6e6e6')
                    yield c

            ws.append(styled_cells(headers))
            for i in ttl_list:
                ws.append(i)

            # Пару пробелов для читаемости
            ws.append([""])
            ws.append([""])
            # Составляем табличку на обувь
            ftw_df = category_df["FTW"]
            ftw_list = ftw_df.values.tolist()
            header = ["FTW"]
            def styled_cells(header):
                for c in header:
                    c = Cell(ws, column="A", row=1, value=c)
                    c.font = Font(bold=True)

                    c.fill = PatternFill(patternType='solid',
                                         fgColor='e6e6e6')
                    yield c
            ws.append(styled_cells(header))

            headers = ['Артикул', "IMG", 'Модель ANTA', 'Сезон', 'RRP', 'Пол',
                       'Коллекция', 'Кол-во\nпродаж', 'Сумма со скидкой', 'Moscow Category\nNet Sales rank']

            def styled_cells(headers):
                for c in headers:
                    c = Cell(ws, column="A", row=1, value=c)
                    c.font = Font(bold=True)

                    c.fill = PatternFill(patternType='solid',
                                         fgColor='e6e6e6')
                    yield c
            ws.append(styled_cells(headers))
            for i in ftw_list:
                ws.append(i)

            # Пару пробелов для читаемости
            ws.append([""])
            ws.append([""])
            # Составляем табличку на одежду
            app_df = category_df["APP"]
            app_list = app_df.values.tolist()
            header = ["APP"]

            def styled_cells(header):
                for c in header:
                    c = Cell(ws, column="A", row=1, value=c)
                    c.font = Font(bold=True)
                    c.fill = PatternFill(patternType='solid',
                                             fgColor='e6e6e6')
                    yield c

            ws.append(styled_cells(header))
            headers = ['Артикул', "IMG", 'Модель ANTA', 'Сезон', 'RRP', 'Пол',
                       'Коллекция', 'Кол-во\nпродаж', 'Сумма со скидкой', 'Moscow Category\nNet Sales rank']

            def styled_cells(headers):
                for c in headers:
                    c = Cell(ws, column="A", row=1, value=c)
                    c.font = Font(bold=True)
                    c.fill = PatternFill(patternType='solid',
                                             fgColor='e6e6e6')
                    yield c
            ws.append(styled_cells(headers))
            for i in app_list:
                ws.append(i)

            # Пару пробелов для читаемости
            ws.append([""])
            ws.append([""])

            # Составляем табличку на аксессуары
            acc_df = category_df["ACC"]
            acc_list = acc_df.values.tolist()
            header = ["ACC"]

            def styled_cells(header):
                for c in header:
                    c = Cell(ws, column="A", row=1, value=c)
                    c.font = Font(bold=True)
                    c.fill = PatternFill(patternType='solid',
                                         fgColor='e6e6e6')
                    yield c

            ws.append(styled_cells(header))
            headers = ['Артикул', "IMG", 'Модель ANTA',  'Сезон', 'RRP', 'Пол',
                       'Коллекция', 'Кол-во\nпродаж', 'Сумма со скидкой', 'Moscow Category\nNet Sales rank']

            def styled_cells(headers):
                for c in headers:
                    c = Cell(ws, column="A", row=1, value=c)
                    c.font = Font(bold=True)
                    c.fill = PatternFill(patternType='solid',
                                         fgColor='e6e6e6')
                    yield c

            ws.append(styled_cells(headers))
            for i in acc_list:
                ws.append(i)

            for row in ws:
                for cell in row:
                    cell.alignment = Alignment(wrapText=True, horizontal='center', vertical='center')

            for cell in ws["E"]:
                if cell.value not in ["", "RRP"]:
                    cell.number_format = "# ##0"

            for cell in ws["I"]:
                if cell.value not in ["", "Сумма со скидкой"]:
                    cell.number_format = "# ##0"

            for col in ws.columns:
                max_length = 0
                column = get_column_letter(col[0].column)  # Get the column name
                for cell in col:
                    try:  # Necessary to avoid error on empty cells
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 4) * 1.2
                ws.column_dimensions[column].width = adjusted_width
            ws.column_dimensions['B'].width = 12
            ws.column_dimensions['H'].width = 12
            ws.column_dimensions['J'].width = 20

            for i in range(1, ws.max_row + 1):
                if ws.cell(i, 4).value != "":
                    rd = ws.row_dimensions[i]
                    rd.height = 50

            row_numb = 1
            for cell in ws['A']:
                if cell.value not in ["Артикул", "APP", "ACC", "FTW", "TOTAL", ""]:
                    self.photo = GetImage().get_photo(cell.value)
                    if self.photo is None:
                        self.photo = GetImage().get_photo("No_photo")
                    try:
                        image_bytes = base64.b64decode(self.photo[0])
                        image_buffer = io.BytesIO(image_bytes)
                        img = openpyxl.drawing.image.Image(image_buffer)
                    except:
                        print(f"Ошибка с {cell.value}")
                        self.photo = GetImage().get_photo("No_photo")
                        image_bytes = base64.b64decode(self.photo[0])
                        image_buffer = io.BytesIO(image_bytes)
                        img = openpyxl.drawing.image.Image(image_buffer)
                    img.height = 60
                    img.width = 70
                    number = ws.cell(row=row_numb, column=2)
                    img.anchor = number.coordinate
                    ws.add_image(img)
                row_numb += 1

        wb.save("Best Sellers Report.xlsx")

class CreateStoreDF:

    def __init__(self):
        self.store_names = None
        self.df_names = ['Магазин', 'Артикул', 'Наименование', 'Количество', 'Категория', 'Сезон', 'SKU', 'RRP',
                         'RRP Amount', 'Grade', 'Пол', 'Коллекция', 'Сумма со скидкой', 'Кол-во продаж', 'Модель ANTA']
        self.is_df_correct = True
        self.store_dataframes = {}

        self.total_ftw_df = None
        self.total_app_df = None
        self.total_acc_df = None
        self.total_ttl_df = None

        self.bs_df = {}

    def create_store_bs(self, main_df):
        """
        Create DF for each category in store

        :param main_df: dict (Dict format {Store: Store DF...}
        :return: dict (Dict format {Store: {Category: {Category DF}...}
        """
        ftw2_df = self.total_ftw_df
        app2_df = self.total_app_df
        acc2_df = self.total_acc_df
        ttl2_df = self.total_ttl_df

        for keys in main_df:
            category = {}
            store_df = main_df[keys]
            # Обработка DF для каждой категории, сортировка по продажам в штучках и вывод первых 10 рез-ов
            ftw1_df = store_df.loc[store_df['Категория'] == "FTW"].sort_values(['Кол-во продаж', 'Сумма со скидкой'],
                                                                              ascending=[False, False]).head(n=10)
            ftw1_df["IMG"] = ""
            app1_df = store_df.loc[store_df['Категория'] == "APP"].sort_values(['Кол-во продаж', 'Сумма со скидкой'],
                                                                              ascending=[False, False]).head(n=10)
            app1_df["IMG"] = ""
            acc1_df = store_df.loc[store_df['Категория'] == "ACC"].sort_values(['Кол-во продаж', 'Сумма со скидкой'],
                                                                              ascending=[False, False]).head(n=10)
            acc1_df["IMG"] = ""
            ttl1_df = store_df.sort_values(['Сумма со скидкой', 'Кол-во продаж'],
                                          ascending=[False, False]).head(n=10)
            ttl1_df["IMG"] = ""

            ttl_df = pd.merge(ttl1_df, ttl2_df[['Артикул', 'Rank by all stores']],
                              on='Артикул', how='left').drop_duplicates(subset=['Артикул'])

            ftw_df = pd.merge(ftw1_df, ftw2_df[['Артикул', 'Rank by all stores']],
                              on='Артикул', how='left').drop_duplicates(subset=['Артикул'])

            app_df = pd.merge(app1_df, app2_df[['Артикул', 'Rank by all stores']],
                              on='Артикул', how='left').drop_duplicates(subset=['Артикул'])

            acc_df = pd.merge(acc1_df, acc2_df[['Артикул', 'Rank by all stores']],
                              on='Артикул', how='left').drop_duplicates(subset=['Артикул'])

            # Удаление из топа продаж результатов с продажами 0
            final_ftw_df = ftw_df.loc[ftw_df['Кол-во продаж'] > 0]
            final_app_df = app_df.loc[app_df['Кол-во продаж'] > 0]
            final_acc_df = acc_df.loc[acc_df['Кол-во продаж'] > 0]
            final_ttl_df = ttl_df.loc[ttl_df['Кол-во продаж'] > 0]
            # Занесение результатов по каждой категории с определенными колонками
            category['FTW'] = final_ftw_df[['Артикул', "IMG", 'Модель ANTA', 'Сезон', 'RRP', 'Пол',
                                            'Коллекция', 'Кол-во продаж', 'Сумма со скидкой', "Rank by all stores"]]
            category['APP'] = final_app_df[['Артикул', "IMG", 'Модель ANTA', 'Сезон', 'RRP', 'Пол',
                                            'Коллекция', 'Кол-во продаж', 'Сумма со скидкой', "Rank by all stores"]]
            category['ACC'] = final_acc_df[['Артикул', "IMG", 'Модель ANTA', 'Сезон', 'RRP', 'Пол',
                                            'Коллекция', 'Кол-во продаж', 'Сумма со скидкой', "Rank by all stores"]]
            category['TTL'] = final_ttl_df[['Артикул', "IMG", 'Модель ANTA', 'Сезон', 'RRP', 'Пол',
                                            'Коллекция', 'Кол-во продаж', 'Сумма со скидкой', "Rank by all stores"]]

            self.bs_df[keys] = category

    def create_df(self, file_name):
        """
        Creation DF to each store for all category

        :param file_name: string (File path)
        :return:  dict (Dict format {Store: Store DF...}
        """
        df = pd.read_excel(file_name)   # Читаем загрузочный файл
        names = df.columns.to_list()  # Получаем заголовки таблицы и далее сравниваем их с необходимыми заголовками
        for header in self.df_names:
            if header not in names:
                self.is_df_correct = False
        if self.is_df_correct:
            self.process_df(df)   # Запускаем обработку файла
        self.create_store_bs(self.store_dataframes)
        return self.bs_df

    def process_df(self, data):
        df = data
        self.store_names = df['Магазин'].unique()   # Получаем уникальные магазины из  файла
        total = df.groupby("Артикул")[['Сумма со скидкой', 'Кол-во продаж']].sum().reset_index()   # Создаем общий DF на все магазины
        total_df = pd.merge(total,
                            df[['Артикул', 'Наименование', 'Модель ANTA', 'Категория', 'Сезон', 'SKU', 'RRP',
                                'Пол', 'Коллекция']],
                            on='Артикул', how='left').drop_duplicates(subset=['Артикул'])


        total_ttl = df.groupby("Артикул")[
            ['Сумма со скидкой', 'Кол-во продаж']].sum().reset_index()  # Создаем общий DF на все магазины
        self.total_ttl_df = pd.merge(total_ttl,
                            df[['Артикул', 'Наименование', 'Модель ANTA', 'Категория', 'Сезон', 'SKU', 'RRP',
                                'Пол', 'Коллекция']],
                            on='Артикул', how='left').drop_duplicates(subset=['Артикул'])
        self.total_ttl_df["Rank by all stores"] = self.total_ttl_df["Сумма со скидкой"].rank(ascending=False,
                                                                           method='first', na_option='bottom')

        total_acc = df.loc[df['Категория'] == "ACC"].groupby("Артикул")[
            ['Сумма со скидкой', 'Кол-во продаж']].sum().reset_index()  # Создаем общий DF на все магазины
        self.total_acc_df = pd.merge(total_acc,
                            df[['Артикул', 'Наименование', 'Модель ANTA', 'Категория', 'Сезон', 'SKU', 'RRP',
                                'Пол', 'Коллекция']],
                            on='Артикул', how='left').drop_duplicates(subset=['Артикул'])
        self.total_acc_df["Rank by all stores"] = self.total_acc_df["Сумма со скидкой"].rank(ascending=False,
                                                                           method='first', na_option='bottom')

        total_app = df.loc[df['Категория'] == "APP"].groupby("Артикул")[
            ['Сумма со скидкой', 'Кол-во продаж']].sum().reset_index()  # Создаем общий DF на все магазины
        self.total_app_df = pd.merge(total_app,
                                     df[['Артикул', 'Наименование', 'Модель ANTA', 'Категория', 'Сезон', 'SKU', 'RRP',
                                         'Пол', 'Коллекция']],
                                     on='Артикул', how='left').drop_duplicates(subset=['Артикул'])
        self.total_app_df["Rank by all stores"] = self.total_app_df["Сумма со скидкой"].rank(ascending=False,
                                                                                             method='first',
                                                                                             na_option='bottom')

        total_ftw = df.loc[df['Категория'] == "FTW"].groupby("Артикул")[
            ['Сумма со скидкой', 'Кол-во продаж']].sum().reset_index()  # Создаем общий DF на все магазины
        self.total_ftw_df = pd.merge(total_ftw,
                                     df[['Артикул', 'Наименование', 'Модель ANTA', 'Категория', 'Сезон', 'SKU', 'RRP',
                                         'Пол', 'Коллекция']],
                                     on='Артикул', how='left').drop_duplicates(subset=['Артикул'])
        self.total_ftw_df["Rank by all stores"] = self.total_ftw_df["Сумма со скидкой"].rank(ascending=False,
                                                                                             method='first',
                                                                                             na_option='bottom')

        self.store_dataframes["Все магазины"] = total_df
        for store in self.store_names:
            store_df = df.loc[df['Магазин'] == store]
            self.store_dataframes[store] = store_df     # Заполняем словарь Ключ(Магазин)-Значение(DF магазина)


if __name__ == "__main__":
    start = Processing()
    start.process("TestFile.xlsx")
