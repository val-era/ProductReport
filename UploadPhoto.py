import sqlite3
import openpyxl
import base64
import io
import os
import pandas as pd
from PIL import JpegImagePlugin


class GetImage:

    def __init__(self):
        self.folder_path = None
        self.file_name = None

    def start_get_img(self):
        self.HoHo()
        self.get_photo_list()  # Запуск любой функции

    def open_folder(self):
        self.folder_path = input(
            "Введите путь к папке с фотографиями. Фото должны быть в формате одна фотография на артикул: \n [+]")
        directory = os.fsencode(self.folder_path)   # Функция записи фотографий в БД
        for file in os.listdir(directory):
            filename = os.fsdecode(file)
            file_name = f"{self.folder_path}\\{filename}"
            with open(file_name, "rb") as image_file:
                photo_base = base64.b64encode(image_file.read())
                art = (filename[: filename.find('.')])
            self.write_sql(art, photo_base)

    def replace_photo(self):
        self.folder_path = input(
            "Введите путь к папке с фотографиями. Фото должны быть в формате одна фотография на артикул: \n [+]")
        directory = os.fsencode(self.folder_path)  # Функция записи фотографий в БД
        for file in os.listdir(directory):
            filename = os.fsdecode(file)
            file_name = f"{self.folder_path}\\{filename}"
            with open(file_name, "rb") as image_file:
                photo_base = base64.b64encode(image_file.read())
                art = (filename[: filename.find('.')])
            self.del_photo(art)
            self.write_sql(art, photo_base)

    def save_photo_infolder(self):
        file_name = input(
            "Введите путь к файлу эксель с артикулами для создания фото, Артикулы должны быть в колонке A:A \n [+]")  # Сохранение фото в папку из БД
        photos = input("Введите путь к папке где должны быть фото \n {+}")
        wb = openpyxl.load_workbook(file_name)
        ws = wb.active
        df_photo = []
        for cell in ws['A']:
            photo = self.get_photo(cell.value)
            if photo is None:
                df_photo.append(cell.value)
            else:
                image_bytes = base64.b64decode(photo[0])
                with open(f'{photos}\\{cell.value}.jpg', 'wb') as img_file:
                    img_file.write(image_bytes)
        df = pd.DataFrame({
            "No_PHOTO": df_photo
        })
        df.to_excel(f'{photos}\\Артикула без фото.xlsx')

    def check_photo_list(self):
        file_name = input("Введите путь к файлу эксель с артикулами для удаления, Артикулы должны быть в колонке A:A \n [+]")   # Проверка фотографий по списку из эксель файла на соответствие
        photos = input("Введите путь к папке с фотографиями \n {+}")
        wb = openpyxl.load_workbook(file_name)
        ws = wb.active
        df_photo = []
        for cell in ws['A']:
            df_photo.append(cell.value)
        directory = os.fsencode(photos)
        for file in os.listdir(directory):
            filename = os.fsdecode(file)
            file_name = f"{photos}\\{filename}"
            art = (filename[: filename.find('.')])
            if art in df_photo:
                pass
            else:
                os.remove(file_name)

    def HoHo(self):
        JpegImagePlugin._getmp = lambda: None
        self.file_hoho_path = input(
            "Введите путь к файлу эксель с артикулами для загрузки фото.\nОбратите внимание что артикулы должны быть в "
            "колонке A:A, Фото грузятся в колонку B:B\nФайл грузится в директорию старого файла (К имени добавляется "
            "NEW) \n [+]")  # Подтягивание фоток в файл эксель

        drive, path = os.path.splitdrive(self.file_hoho_path)
        path, filename = os.path.split(path)

        wb = openpyxl.load_workbook(self.file_hoho_path)
        ws = wb.active
        ws.column_dimensions['B'].width = 12

        for i in range(1, ws.max_row + 1):
            if ws.cell(i, 4).value != "":
                rd = ws.row_dimensions[i]
                rd.height = 50

        row_numb = 1
        for cell in ws['A']:
            self.photo = self.get_photo(cell.value)
            if self.photo is None:
                self.photo = self.get_photo("No_photo")
            image_bytes = base64.b64decode(self.photo[0])
            image_buffer = io.BytesIO(image_bytes)
            img = openpyxl.drawing.image.Image(image_buffer)
            img.height = 60
            img.width = 70
            number = ws.cell(row=row_numb, column=2)
            img.anchor = number.coordinate
            ws.add_image(img)
            row_numb += 1
        wb.save(f"{path}\\NEW{filename}")

    def write_sql(self, art, photo):
        con = sqlite3.connect("ProductReportDB.db")   # Запись в базу данных фотографий в формате строки для кодировки
        cursor = con.cursor()
        cursor.execute('''
        CREATE TABLE IF NOT EXISTS Photos (
        art TEXT,
        photo TEXT
        )
        ''')
        cursor.execute('INSERT INTO Photos (art, photo) VALUES (?, ?)',
                       (art, photo))
        con.commit()
        con.close()

    def get_photo(self, art):
        con = sqlite3.connect("ProductReportDB.db")    # получение фото из базы данных
        cursor = con.cursor()
        cursor.execute('''
                CREATE TABLE IF NOT EXISTS Photos (
                art TEXT,
                photo TEXT
                )
                ''')
        cursor.execute('SELECT photo FROM Photos WHERE art=?', [art])
        photo = cursor.fetchone()
        con.commit()
        con.close()
        return photo

    def get_photo_list(self):
        con = sqlite3.connect("ProductReportDB.db")   # Получение листа с фотографиями
        cursor = con.cursor()
        cursor.execute('''
                        CREATE TABLE IF NOT EXISTS Photos (
                        art TEXT,
                        photo TEXT
                        )
                        ''')
        cursor.execute('SELECT art FROM Photos')
        photo = cursor.fetchall()
        con.commit()
        con.close()
        art_list = []
        for i in photo:
            if i not in art_list:
                art_list.append(i[0])
        df = pd.DataFrame({
            "Артикул с фото": art_list})
        df.to_excel('Фото из БД.xlsx')

    def del_photo_file(self):
        file_name = input("Введите путь к файлу эксель с артикулами для удаления. Артикулы должны быть в колонке A:A \n [+]")    # Удаление из БД по файлу
        df = pd.read_excel(file_name)
        for index, row in df.iterrows():
            for i in row.values:
                self.del_photo(i)

    def del_photo(self, art):
        con = sqlite3.connect("ProductReportDB.db")   # Удаление по артикулу
        cursor = con.cursor()
        cursor.execute('''
                        CREATE TABLE IF NOT EXISTS Photos (
                        art TEXT,
                        photo TEXT
                        )
                        ''')
        cursor.execute('DELETE FROM Photos WHERE art=?', [art])
        con.commit()
        con.close()


if __name__ == "__main__":
    start = GetImage()
    start.start_get_img()
