import base64
import io

import numpy
import pandas as pd
import openpyxl
from openpyxl.cell.cell import Cell
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from UploadPhoto import GetImage
from PIL import Image


class Processing:
    def __init__(self):
        self.store_by_wmn_cat = None
        self.store_by_men_cat = None
        self.store_by_category = None
        self.store_by_gender = None
        self.file_name = "TestFile.xlsx"

    def process(self, file_name):
        self.file_name = file_name
        start_cr_df = CreateStoreDF()  # Класс чтения файла и разбивки на DF магазина
        (self.store_by_gender, self.store_by_category,
         self.store_by_men_cat, self.store_by_wmn_cat,
         self.store_by_app_coll, self.store_by_ftw_coll,
         self.store_by_product_type, self.store_by_models,
         self.store_by_kids_cat, self.store_by_app_gender, self.store_by_ftw_gender) = start_cr_df.read_file(self.file_name)
        self.excel_file()

    def excel_file(self):
        wb = openpyxl.Workbook()
        worksheet = wb.active
        wb.remove(worksheet)

        ws = wb.create_sheet(title="Информация и метрики")

        image_bytes = base64.b64decode(GetImage().get_photo("Information")[0])

        image_buffer = io.BytesIO(image_bytes)
        img = openpyxl.drawing.image.Image(image_buffer)
        img.height = 654
        img.width = 1024
        number = ws.cell(row=1, column=1)
        img.anchor = number.coordinate
        ws.add_image(img)

        for store in self.store_by_gender:
            wb.create_sheet(title=store)
            gender_df = self.store_by_gender[store]
            category_df = self.store_by_category[store]
            men_cat_df = self.store_by_men_cat[store]
            wmn_cat_df = self.store_by_wmn_cat[store]
            app_coll_df = self.store_by_app_coll[store]
            ftw_coll_df = self.store_by_ftw_coll[store]
            pr_name_q = self.store_by_product_type[store]
            models_q = self.store_by_models[store]
            kids_cat_df = self.store_by_kids_cat[store]
            by_app_gender = self.store_by_app_gender[store]
            by_ftw_gender = self.store_by_ftw_gender[store]
            ws = wb[store]
            ws.alignment = Alignment(wrapText=True)

            pr_name = sorted(pr_name_q, key=lambda x: x.values[0][1], reverse=True)
            models = sorted(models_q, key=lambda x: x.values[0][1], reverse=True)

            header = ["BY GENDERS"]

            def styled_cells(header):
                for c in header:
                    c = Cell(ws, column="A", row=1, value=c)
                    c.font = Font(bold=True)
                    c.fill = PatternFill(patternType='solid',
                                         fgColor='e6e6e6')
                    yield c

            ws.append(styled_cells(header))
            headers = ['GENDER', "NS", 'NS MIX %', 'AVERAGE NS MIX %', 'NS QTY', 'NS QTY MIX %', 'STOCK QTY',
                       'STOCK QTY MIX %', 'SKU', 'SKU MIX %', 'GRADE A MIX %', "GRADE B MIX %", "GRADE C MIX %", "GRADE C SKU"]

            def styled_cells(headers):
                for c in headers:
                    c = Cell(ws, column="A", row=1, value=c)
                    c.font = Font(bold=True)
                    c.fill = PatternFill(patternType='solid',
                                         fgColor='e6e6e6')
                    yield c

            ws.append(styled_cells(headers))
            for i in gender_df:
                for v in i.values.tolist():
                    ws.append(v)

            # Пару пробелов для читаемости
            ws.append([""])
            ws.append([""])

            header = ["APP BY GENDERS"]

            def styled_cells(header):
                for c in header:
                    c = Cell(ws, column="A", row=1, value=c)
                    c.font = Font(bold=True)
                    c.fill = PatternFill(patternType='solid',
                                         fgColor='e6e6e6')
                    yield c

            ws.append(styled_cells(header))
            headers = ['GENDER', "NS", 'NS MIX %', 'AVERAGE NS MIX %', 'NS QTY', 'NS QTY MIX %', 'STOCK QTY',
                       'STOCK QTY MIX %', 'SKU', 'SKU MIX %', 'GRADE A MIX %', "GRADE B MIX %", "GRADE C MIX %",
                       "GRADE C SKU"]

            def styled_cells(headers):
                for c in headers:
                    c = Cell(ws, column="A", row=1, value=c)
                    c.font = Font(bold=True)
                    c.fill = PatternFill(patternType='solid',
                                         fgColor='e6e6e6')
                    yield c

            ws.append(styled_cells(headers))
            for i in by_app_gender:
                for v in i.values.tolist():
                    ws.append(v)

            # Пару пробелов для читаемости
            ws.append([""])
            ws.append([""])

            header = ["FTW BY GENDERS"]

            def styled_cells(header):
                for c in header:
                    c = Cell(ws, column="A", row=1, value=c)
                    c.font = Font(bold=True)
                    c.fill = PatternFill(patternType='solid',
                                         fgColor='e6e6e6')
                    yield c

            ws.append(styled_cells(header))
            headers = ['GENDER', "NS", 'NS MIX %', 'AVERAGE NS MIX %', 'NS QTY', 'NS QTY MIX %', 'STOCK QTY',
                       'STOCK QTY MIX %', 'SKU', 'SKU MIX %', 'GRADE A MIX %', "GRADE B MIX %", "GRADE C MIX %",
                       "GRADE C SKU"]

            def styled_cells(headers):
                for c in headers:
                    c = Cell(ws, column="A", row=1, value=c)
                    c.font = Font(bold=True)
                    c.fill = PatternFill(patternType='solid',
                                         fgColor='e6e6e6')
                    yield c

            ws.append(styled_cells(headers))
            for i in by_ftw_gender:
                for v in i.values.tolist():
                    ws.append(v)

            # Пару пробелов для читаемости
            ws.append([""])
            ws.append([""])

            ws.append([""])
            ws.append([""])

            header = ["BY CATEGORY ALL GENDERS"]

            def styled_cells(header):
                for c in header:
                    c = Cell(ws, column="A", row=1, value=c)
                    c.font = Font(bold=True)
                    c.fill = PatternFill(patternType='solid',
                                         fgColor='e6e6e6')
                    yield c

            ws.append(styled_cells(header))
            headers = ['CATEGORY', "NS", 'NS MIX %', 'AVERAGE NS MIX %', 'NS QTY', 'NS QTY MIX %', 'STOCK QTY',
                       'STOCK QTY MIX %', 'SKU','SKU MIX %', 'GRADE A MIX %', "GRADE B MIX %", "GRADE C MIX %", "GRADE C SKU"]

            def styled_cells(headers):
                for c in headers:
                    c = Cell(ws, column="A", row=1, value=c)
                    c.font = Font(bold=True)
                    c.fill = PatternFill(patternType='solid',
                                         fgColor='e6e6e6')
                    yield c

            ws.append(styled_cells(headers))
            for i in category_df:
                for v in i.values.tolist():
                    ws.append(v)

            # Пару пробелов для читаемости
            ws.append([""])
            ws.append([""])

            header = ["BY CATEGORY MEN"]

            def styled_cells(header):
                for c in header:
                    c = Cell(ws, column="A", row=1, value=c)
                    c.font = Font(bold=True)
                    c.fill = PatternFill(patternType='solid',
                                         fgColor='e6e6e6')
                    yield c

            ws.append(styled_cells(header))
            headers = ['CATEGORY', "NS", 'NS MIX %', 'AVERAGE NS MIX %', 'NS QTY', 'NS QTY MIX %', 'STOCK QTY',
                       'STOCK QTY MIX %', 'SKU','SKU MIX %', 'GRADE A MIX %', "GRADE B MIX %", "GRADE C MIX %", "GRADE C SKU"]

            def styled_cells(headers):
                for c in headers:
                    c = Cell(ws, column="A", row=1, value=c)
                    c.font = Font(bold=True)
                    c.fill = PatternFill(patternType='solid',
                                         fgColor='e6e6e6')
                    yield c

            ws.append(styled_cells(headers))
            for i in men_cat_df:
                for v in i.values.tolist():
                    ws.append(v)

            ws.append([""])
            ws.append([""])

            header = ["BY CATEGORY WOMEN"]

            def styled_cells(header):
                for c in header:
                    c = Cell(ws, column="A", row=1, value=c)
                    c.font = Font(bold=True)
                    c.fill = PatternFill(patternType='solid',
                                         fgColor='e6e6e6')
                    yield c

            ws.append(styled_cells(header))
            headers = ['CATEGORY', "NS", 'NS MIX %', 'AVERAGE NS MIX %', 'NS QTY', 'NS QTY MIX %', 'STOCK QTY',
                       'STOCK QTY MIX %', 'SKU','SKU MIX %', 'GRADE A MIX %', "GRADE B MIX %", "GRADE C MIX %", "GRADE C SKU"]

            def styled_cells(headers):
                for c in headers:
                    c = Cell(ws, column="A", row=1, value=c)
                    c.font = Font(bold=True)
                    c.fill = PatternFill(patternType='solid',
                                         fgColor='e6e6e6')
                    yield c

            ws.append(styled_cells(headers))
            for i in wmn_cat_df:
                for v in i.values.tolist():
                    ws.append(v)

            if kids_cat_df[-1].values[0][8] > 0:
                ws.append([""])
                ws.append([""])

                header = ["BY CATEGORY KIDS"]
                def styled_cells(header):
                    for c in header:
                        c = Cell(ws, column="A", row=1, value=c)
                        c.font = Font(bold=True)
                        c.fill = PatternFill(patternType='solid',
                                             fgColor='e6e6e6')
                        yield c

                ws.append(styled_cells(header))
                headers = ['CATEGORY', "NS", 'NS MIX %', 'AVERAGE NS MIX %', 'NS QTY', 'NS QTY MIX %', 'STOCK QTY',
                           'STOCK QTY MIX %', 'SKU', 'SKU MIX %', 'GRADE A MIX %', "GRADE B MIX %", "GRADE C MIX %",
                           "GRADE C SKU"]

                def styled_cells(headers):
                    for c in headers:
                        c = Cell(ws, column="A", row=1, value=c)
                        c.font = Font(bold=True)
                        c.fill = PatternFill(patternType='solid',
                                             fgColor='e6e6e6')
                        yield c

                ws.append(styled_cells(headers))
                for i in kids_cat_df:
                    for v in i.values.tolist():
                        ws.append(v)

            # Пару пробелов для читаемости
            ws.append([""])
            ws.append([""])

            # Пару пробелов для читаемости
            ws.append([""])
            ws.append([""])

            header = ["APP BY COLLECTIONS"]

            def styled_cells(header):
                for c in header:
                    c = Cell(ws, column="A", row=1, value=c)
                    c.font = Font(bold=True)
                    c.fill = PatternFill(patternType='solid',
                                         fgColor='e6e6e6')
                    yield c

            ws.append(styled_cells(header))
            headers = ['COLLECTIONS', "NS", 'NS MIX %', 'AVERAGE NS MIX %', 'NS QTY', 'NS QTY MIX %', 'STOCK QTY',
                       'STOCK QTY MIX %', 'SKU','SKU MIX %', 'GRADE A MIX %', "GRADE B MIX %", "GRADE C MIX %", "GRADE C SKU"]

            def styled_cells(headers):
                for c in headers:
                    c = Cell(ws, column="A", row=1, value=c)
                    c.font = Font(bold=True)
                    c.fill = PatternFill(patternType='solid',
                                         fgColor='e6e6e6')
                    yield c

            ws.append(styled_cells(headers))
            for i in app_coll_df:
                if not i.isnull().any().values[0]:
                    for v in i.values.tolist():
                        ws.append(v)
            ws.append([""])
            ws.append([""])

            header = ["FTW BY COLLECTIONS"]

            def styled_cells(header):
                for c in header:
                    c = Cell(ws, column="A", row=1, value=c)
                    c.font = Font(bold=True)
                    c.fill = PatternFill(patternType='solid',
                                         fgColor='e6e6e6')
                    yield c

            ws.append(styled_cells(header))
            headers = ['COLLECTIONS', "NS", 'NS MIX %', 'AVERAGE NS MIX %', 'NS QTY', 'NS QTY MIX %', 'STOCK QTY',
                       'STOCK QTY MIX %', 'SKU','SKU MIX %', 'GRADE A MIX %', "GRADE B MIX %", "GRADE C MIX %", "GRADE C SKU"]

            def styled_cells(headers):
                for c in headers:
                    c = Cell(ws, column="A", row=1, value=c)
                    c.font = Font(bold=True)
                    c.fill = PatternFill(patternType='solid',
                                         fgColor='e6e6e6')
                    yield c

            ws.append(styled_cells(headers))
            for i in ftw_coll_df:
                if not i.isnull().any().values[0]:
                    for v in i.values.tolist():
                        ws.append(v)

            ws.append([""])
            ws.append([""])

            # Пару пробелов для читаемости
            ws.append([""])
            ws.append([""])

            header = ["BY PRODUCT TYPES"]

            def styled_cells(header):
                for c in header:
                    c = Cell(ws, column="A", row=1, value=c)
                    c.font = Font(bold=True)
                    c.fill = PatternFill(patternType='solid',
                                         fgColor='e6e6e6')
                    yield c

            ws.append(styled_cells(header))
            headers = ['COLLECTIONS', "NS", 'NS MIX %', 'AVERAGE NS MIX %', 'NS QTY', 'NS QTY MIX %', 'STOCK QTY',
                       'STOCK QTY MIX %', 'SKU', 'SKU MIX %']

            def styled_cells(headers):
                for c in headers:
                    c = Cell(ws, column="A", row=1, value=c)
                    c.font = Font(bold=True)
                    c.fill = PatternFill(patternType='solid',
                                         fgColor='e6e6e6')
                    yield c

            ws.append(styled_cells(headers))
            for i in pr_name:
                if i.values[0][1] > 0 and i.values[0][0] != "TTL":
                    if not i.isnull().any().values[0]:
                        for v in i.values.tolist():
                            ws.append(v)

            ws.append([""])
            ws.append([""])

            header = ["BY MODELS NAME"]

            def styled_cells(header):
                for c in header:
                    c = Cell(ws, column="A", row=1, value=c)
                    c.font = Font(bold=True)
                    c.fill = PatternFill(patternType='solid',
                                         fgColor='e6e6e6')
                    yield c

            ws.append(styled_cells(header))
            headers = ['COLLECTIONS', "NS", 'NS MIX %', 'AVERAGE NS MIX %', 'NS QTY', 'NS QTY MIX %', 'STOCK QTY',
                       'STOCK QTY MIX %', 'SKU', 'SKU MIX %']

            def styled_cells(headers):
                for c in headers:
                    c = Cell(ws, column="A", row=1, value=c)
                    c.font = Font(bold=True)
                    c.fill = PatternFill(patternType='solid',
                                         fgColor='e6e6e6')
                    yield c

            ws.append(styled_cells(headers))
            for i in models:
                if i.values[0][1] > 0 and i.values[0][0] != "TTL":
                    if not i.isnull().any().values[0]:
                        for v in i.values.tolist():
                            ws.append(v)

            row_numb = 1
            rows_ttl = []
            for cell in ws["A"]:
                if cell.value == "TTL":
                    rows_ttl.append(row_numb)
                row_numb += 1

            for i in rows_ttl:
                for row in ws.iter_rows(min_row=i, max_col=None, max_row=i):
                    for cell in row:
                        if cell.value != "":
                            cell.font = Font(bold=True)
                            cell.fill = PatternFill(patternType='solid',
                                                    fgColor='cfcfcf')

            for row in ws:
                for cell in row:
                    cell.alignment = Alignment(wrapText=True, horizontal='center', vertical='center')

            num = ["B", "E", "G", "I", "N"]
            pers = ["C", "D", "F", "H", "J", "K", "L", "M"]

            for i in num:
                for cell in ws[i]:
                    cell.number_format = "### ### ##0"

            for i in pers:
                for cell in ws[i]:
                    cell.number_format = '0%'

            ws.column_dimensions['A'].width = 40
            ws.column_dimensions['B'].width = 20
            ws.column_dimensions['D'].width = 12
            ws.column_dimensions['H'].width = 12
        sheet = wb["Все магазины"]
        col = ["D", "H", "L"]
        for column in col:
            which_cols = openpyxl.utils.cell.column_index_from_string(column)
            sheet.delete_cols(which_cols, 1)
        wb.save("Sales Analysis Report.xlsx")


class CreateStoreDF:
    def __init__(self):
        self.store_name = None
        self.path = None
        self.df_names = ['Магазин', 'Артикул', 'Наименование', 'Количество', 'Категория', 'Сезон', 'SKU', 'RRP',
                         'RRP Amount', 'Grade', 'Пол', 'Коллекция', 'Сумма со скидкой', 'Кол-во продаж',
                         'Актуальный заказ', 'Тип продукта']
        self.is_df_correct = True

        self.store_by_gender = {}
        self.store_by_category = {}
        self.store_by_men_cat = {}
        self.store_by_wmn_cat = {}
        self.store_by_app_coll = {}
        self.store_by_ftw_coll = {}
        self.store_by_product_type = {}
        self.store_by_models = {}
        self.store_by_kids_cat = {}
        self.store_by_ftw_gender = {}
        self.store_by_app_gender = {}

    def read_file(self, path):
        """
        Creation DF to each store for all category

        :param path: string (File path)
        :return:  dict (Dict format {Store: Store DF...}
        """
        self.path = path
        df = pd.read_excel(self.path)  # Читаем загрузочный файл
        names = df.columns.to_list()  # Получаем заголовки таблицы и далее сравниваем их с необходимыми заголовками
        for header in self.df_names:
            if header not in names:
                self.is_df_correct = False
        if self.is_df_correct:
            self.process_df(df)  # Запускаем обработку файла
            return (self.store_by_gender, self.store_by_category, self.store_by_men_cat,
                    self.store_by_wmn_cat, self.store_by_app_coll, self.store_by_ftw_coll,
                    self.store_by_product_type, self.store_by_models, self.store_by_kids_cat,
                    self.store_by_app_gender, self.store_by_ftw_gender)

    def process_df(self, df):
        numpy.seterr(divide='ignore', invalid='ignore')
        self.get_gender_df(df)
        self.get_category_df(df)
        self.get_men_cat(df)
        self.get_wmn_cat(df)
        self.get_app_collect(df)
        self.get_ftw_collect(df)
        self.get_pr_type(df)
        self.get_by_model(df)
        self.get_kids_cat(df)
        self.get_ftw_gender(df)
        self.get_app_gender(df)

    def get_app_gender(self, df_1):
        numpy.seterr(divide='ignore', invalid='ignore')
        df = df_1.loc[df_1['Категория'] == "APP"]
        self.store_name = df['Магазин'].unique()
        self.collections_name = df['Пол'].unique()

        cat_df = []
        for category in self.collections_name:  # Тотал по всем магазинам
            category_df = df.loc[(df['Пол'] == category) & (df['Категория'] == "APP")]
            ns = category_df['Сумма со скидкой'].sum()
            ns_mix = category_df['Сумма со скидкой'].sum() / df.loc[df['Категория'] == "APP"]['Сумма со скидкой'].sum()
            ns_qty = category_df['Кол-во продаж'].sum()
            ns_qty_mix = category_df['Кол-во продаж'].sum() / df.loc[df['Категория'] == "APP"]['Кол-во продаж'].sum()
            stock_qty = category_df['Количество'].sum()
            store_qty_mix = category_df['Количество'].sum() / df.loc[df['Категория'] == "APP"]['Количество'].sum()
            sku_mix = category_df.loc[category_df['Актуальный заказ'] == "Актуально"]['SKU'].sum()/df.loc[df['Актуальный заказ'] == "Актуально"]['SKU'].sum()
            act_sku = category_df.loc[category_df['Актуальный заказ'] == "Актуально"]['SKU'].sum()
            sku_a_mix = category_df.loc[(category_df['Grade'] == "A") & (category_df['Актуальный заказ'] == "Актуально")]['SKU'].sum() / act_sku
            sku_b_mix = category_df.loc[(category_df['Grade'] == "B") & (category_df['Актуальный заказ'] == "Актуально")]['SKU'].sum() / act_sku
            sku_c_mix = category_df.loc[(category_df['Grade'] == "C") & (category_df['Актуальный заказ'] == "Актуально")]['SKU'].sum() / act_sku
            sku_c = category_df.loc[(category_df['Grade'] == "C") & (category_df['Актуальный заказ'] == "Актуально")]['SKU'].sum()

            df_ttl_category_ttl = pd.DataFrame({
                "Category": [category],
                "NS": [ns],
                "NS MIX": [ns_mix],
                "NS RD MIX": [ns_mix],
                "NS QTY": [ns_qty],
                "NS QTY MIX": [ns_qty_mix],
                "STOCK QTY": [stock_qty],
                "STOCK QTY MIX": [store_qty_mix],
                "SKU": [act_sku],
                "SKU MIX": [sku_mix],
                "SKU A MIX": [sku_a_mix],
                "SKU B MIX": [sku_b_mix],
                "SKU C MIX": [sku_c_mix],
                "SKU C": [sku_c]
            })
            cat_df.append(df_ttl_category_ttl)

        self.store_by_app_gender["Все магазины"] = cat_df

        for store in self.store_name:
            cat_store_df = []
            store_df = df.loc[(df['Магазин'] == store) & (df['Категория'] == "APP")]

            ns = store_df['Сумма со скидкой'].sum()
            ns_mix = 1
            ns_qty = store_df['Кол-во продаж'].sum()
            ns_qty_mix = 1
            stock_qty = store_df['Количество'].sum()
            store_qty_mix = 1
            sku_mix = 1
            act_sku = store_df.loc[store_df['Актуальный заказ'] == "Актуально"]['SKU'].sum()
            sku_a_mix = store_df.loc[(store_df['Grade'] == "A") & (store_df['Актуальный заказ'] == "Актуально")]['SKU'].sum() / act_sku
            sku_b_mix = store_df.loc[(store_df['Grade'] == "B") & (store_df['Актуальный заказ'] == "Актуально")]['SKU'].sum() / act_sku
            sku_c_mix = store_df.loc[(store_df['Grade'] == "C") & (store_df['Актуальный заказ'] == "Актуально")]['SKU'].sum() / act_sku
            sku_c = store_df.loc[(store_df['Grade'] == "C") & (store_df['Актуальный заказ'] == "Актуально")]['SKU'].sum()

            df_cat_ttl = pd.DataFrame({
                "Category": ["TTL"],
                "NS": [ns],
                "NS MIX": [ns_mix],
                "NS RD MIX": [ns_mix],
                "NS QTY": [ns_qty],
                "NS QTY MIX": [ns_qty_mix],
                "STOCK QTY": [stock_qty],
                "STOCK QTY MIX": [store_qty_mix],
                "SKU": [act_sku],
                "SKU MIX": [sku_mix],
                "SKU A MIX": [sku_a_mix],
                "SKU B MIX": [sku_b_mix],
                "SKU C MIX": [sku_c_mix],
                "SKU C": [sku_c]
            })
            self.collections_name = df.loc[(df['Магазин'] == store) & (df['Категория'] == "APP")]['Пол'].unique()
            for category in self.collections_name:
                category_df = store_df.loc[(store_df['Пол'] == category) & (df['Категория'] == "APP")]
                category1_df = df.loc[(df['Пол'] == category) & (df['Категория'] == "APP")]
                ns_rd_mix = category1_df['Сумма со скидкой'].sum() / df.loc[df['Категория'] == "APP"][
                    'Сумма со скидкой'].sum()
                ns = category_df['Сумма со скидкой'].sum()
                ns_mix = category_df['Сумма со скидкой'].sum() / store_df['Сумма со скидкой'].sum()
                ns_qty = category_df['Кол-во продаж'].sum()
                ns_qty_mix = category_df['Кол-во продаж'].sum() / store_df['Кол-во продаж'].sum()
                stock_qty = category_df['Количество'].sum()
                store_qty_mix = category_df['Количество'].sum() / store_df['Количество'].sum()
                sku_mix = category_df.loc[category_df['Актуальный заказ'] == "Актуально"]['SKU'].sum()/store_df.loc[store_df['Актуальный заказ'] == "Актуально"]['SKU'].sum()
                act_sku = category_df.loc[category_df['Актуальный заказ'] == "Актуально"]['SKU'].sum()
                sku_a_mix = category_df.loc[(category_df['Grade'] == "A") & (category_df['Актуальный заказ'] == "Актуально")]['SKU'].sum() / act_sku
                sku_b_mix = category_df.loc[(category_df['Grade'] == "B") & (category_df['Актуальный заказ'] == "Актуально")]['SKU'].sum() / act_sku
                sku_c_mix = category_df.loc[(category_df['Grade'] == "C") & (category_df['Актуальный заказ'] == "Актуально")]['SKU'].sum() / act_sku
                sku_c = category_df.loc[(category_df['Grade'] == "C") & (category_df['Актуальный заказ'] == "Актуально")]['SKU'].sum()

                df_gender_bystore = pd.DataFrame({
                    "Category": [category],
                    "NS": [ns],
                    "NS MIX": [ns_mix],
                    "NS RD MIX": [ns_rd_mix],
                    "NS QTY": [ns_qty],
                    "NS QTY MIX": [ns_qty_mix],
                    "STOCK QTY": [stock_qty],
                    "STOCK QTY MIX": [store_qty_mix],
                    "SKU": [act_sku],
                    "SKU MIX": [sku_mix],
                    "SKU A MIX": [sku_a_mix],
                    "SKU B MIX": [sku_b_mix],
                    "SKU C MIX": [sku_c_mix],
                    "SKU C": [sku_c]
                })
                cat_store_df.append(df_gender_bystore)
            cat_store_df.append(df_cat_ttl)
            self.store_by_app_gender[store] = cat_store_df

    def get_ftw_gender(self, df_1):
        numpy.seterr(divide='ignore', invalid='ignore')
        df = df_1.loc[df_1['Категория'] == "FTW"]
        self.store_name = df['Магазин'].unique()
        self.collections_name = df['Пол'].unique()

        cat_df = []
        for category in self.collections_name:  # Тотал по всем магазинам
            category_df = df.loc[(df['Пол'] == category) & (df['Категория'] == "FTW")]
            ns = category_df['Сумма со скидкой'].sum()
            ns_mix = category_df['Сумма со скидкой'].sum() / df.loc[df['Категория'] == "FTW"]['Сумма со скидкой'].sum()
            ns_qty = category_df['Кол-во продаж'].sum()
            ns_qty_mix = category_df['Кол-во продаж'].sum() / df.loc[df['Категория'] == "FTW"]['Кол-во продаж'].sum()
            stock_qty = category_df['Количество'].sum()
            store_qty_mix = category_df['Количество'].sum() / df.loc[df['Категория'] == "FTW"]['Количество'].sum()
            sku_mix = category_df.loc[category_df['Актуальный заказ'] == "Актуально"]['SKU'].sum()/df.loc[df['Актуальный заказ'] == "Актуально"]['SKU'].sum()
            act_sku = category_df.loc[category_df['Актуальный заказ'] == "Актуально"]['SKU'].sum()
            sku_a_mix = category_df.loc[(category_df['Grade'] == "A") & (category_df['Актуальный заказ'] == "Актуально")]['SKU'].sum() / act_sku
            sku_b_mix = category_df.loc[(category_df['Grade'] == "B") & (category_df['Актуальный заказ'] == "Актуально")]['SKU'].sum() / act_sku
            sku_c_mix = category_df.loc[(category_df['Grade'] == "C") & (category_df['Актуальный заказ'] == "Актуально")]['SKU'].sum() / act_sku
            sku_c = category_df.loc[(category_df['Grade'] == "C") & (category_df['Актуальный заказ'] == "Актуально")]['SKU'].sum()

            df_ttl_category_ttl = pd.DataFrame({
                "Category": [category],
                "NS": [ns],
                "NS MIX": [ns_mix],
                "NS RD MIX": [ns_mix],
                "NS QTY": [ns_qty],
                "NS QTY MIX": [ns_qty_mix],
                "STOCK QTY": [stock_qty],
                "STOCK QTY MIX": [store_qty_mix],
                "SKU": [act_sku],
                "SKU MIX": [sku_mix],
                "SKU A MIX": [sku_a_mix],
                "SKU B MIX": [sku_b_mix],
                "SKU C MIX": [sku_c_mix],
                "SKU C": [sku_c]
            })
            cat_df.append(df_ttl_category_ttl)

        self.store_by_ftw_gender["Все магазины"] = cat_df

        for store in self.store_name:
            cat_store_df = []
            store_df = df.loc[(df['Магазин'] == store) & (df['Категория'] == "FTW")]

            ns = store_df['Сумма со скидкой'].sum()
            ns_mix = 1
            ns_qty = store_df['Кол-во продаж'].sum()
            ns_qty_mix = 1
            stock_qty = store_df['Количество'].sum()
            store_qty_mix = 1
            sku_mix = 1
            act_sku = store_df.loc[store_df['Актуальный заказ'] == "Актуально"]['SKU'].sum()
            sku_a_mix = store_df.loc[(store_df['Grade'] == "A") & (store_df['Актуальный заказ'] == "Актуально")]['SKU'].sum() / act_sku
            sku_b_mix = store_df.loc[(store_df['Grade'] == "B") & (store_df['Актуальный заказ'] == "Актуально")]['SKU'].sum() / act_sku
            sku_c_mix = store_df.loc[(store_df['Grade'] == "C") & (store_df['Актуальный заказ'] == "Актуально")]['SKU'].sum() / act_sku
            sku_c = store_df.loc[(store_df['Grade'] == "C") & (store_df['Актуальный заказ'] == "Актуально")]['SKU'].sum()

            df_cat_ttl = pd.DataFrame({
                "Category": ["TTL"],
                "NS": [ns],
                "NS MIX": [ns_mix],
                "NS RD MIX": [ns_mix],
                "NS QTY": [ns_qty],
                "NS QTY MIX": [ns_qty_mix],
                "STOCK QTY": [stock_qty],
                "STOCK QTY MIX": [store_qty_mix],
                "SKU": [act_sku],
                "SKU MIX": [sku_mix],
                "SKU A MIX": [sku_a_mix],
                "SKU B MIX": [sku_b_mix],
                "SKU C MIX": [sku_c_mix],
                "SKU C": [sku_c]
            })
            self.collections_name = df.loc[(df['Магазин'] == store) & (df['Категория'] == "FTW")]['Пол'].unique()
            for category in self.collections_name:
                category_df = store_df.loc[(store_df['Пол'] == category) & (df['Категория'] == "FTW")]
                category1_df = df.loc[(df['Пол'] == category) & (df['Категория'] == "FTW")]
                ns_rd_mix = category1_df['Сумма со скидкой'].sum() / df.loc[df['Категория'] == "FTW"][
                    'Сумма со скидкой'].sum()
                ns = category_df['Сумма со скидкой'].sum()
                ns_mix = category_df['Сумма со скидкой'].sum() / store_df['Сумма со скидкой'].sum()
                ns_qty = category_df['Кол-во продаж'].sum()
                ns_qty_mix = category_df['Кол-во продаж'].sum() / store_df['Кол-во продаж'].sum()
                stock_qty = category_df['Количество'].sum()
                store_qty_mix = category_df['Количество'].sum() / store_df['Количество'].sum()
                sku_mix = category_df.loc[category_df['Актуальный заказ'] == "Актуально"]['SKU'].sum()/store_df.loc[store_df['Актуальный заказ'] == "Актуально"]['SKU'].sum()
                act_sku = category_df.loc[category_df['Актуальный заказ'] == "Актуально"]['SKU'].sum()
                sku_a_mix = category_df.loc[(category_df['Grade'] == "A") & (category_df['Актуальный заказ'] == "Актуально")]['SKU'].sum() / act_sku
                sku_b_mix = category_df.loc[(category_df['Grade'] == "B") & (category_df['Актуальный заказ'] == "Актуально")]['SKU'].sum() / act_sku
                sku_c_mix = category_df.loc[(category_df['Grade'] == "C") & (category_df['Актуальный заказ'] == "Актуально")]['SKU'].sum() / act_sku
                sku_c = category_df.loc[(category_df['Grade'] == "C") & (category_df['Актуальный заказ'] == "Актуально")]['SKU'].sum()

                df_gender_bystore = pd.DataFrame({
                    "Category": [category],
                    "NS": [ns],
                    "NS MIX": [ns_mix],
                    "NS RD MIX": [ns_rd_mix],
                    "NS QTY": [ns_qty],
                    "NS QTY MIX": [ns_qty_mix],
                    "STOCK QTY": [stock_qty],
                    "STOCK QTY MIX": [store_qty_mix],
                    "SKU": [act_sku],
                    "SKU MIX": [sku_mix],
                    "SKU A MIX": [sku_a_mix],
                    "SKU B MIX": [sku_b_mix],
                    "SKU C MIX": [sku_c_mix],
                    "SKU C": [sku_c]
                })
                cat_store_df.append(df_gender_bystore)
            cat_store_df.append(df_cat_ttl)
            self.store_by_ftw_gender[store] = cat_store_df



    def get_kids_cat(self, df):
        self.store_name = df['Магазин'].unique()
        self.category_name = df['Категория'].unique()

        cat_df = []
        for category in self.category_name:  # Тотал по всем магазинам
            category_df = df.loc[(df['Категория'] == category) & (df['Пол'] == "KIDS")]
            ns = category_df['Сумма со скидкой'].sum()
            ns_mix = category_df['Сумма со скидкой'].sum() / df.loc[df['Пол'] == "KIDS"]['Сумма со скидкой'].sum()
            ns_qty = category_df['Кол-во продаж'].sum()
            ns_qty_mix = category_df['Кол-во продаж'].sum() / df.loc[df['Пол'] == "KIDS"]['Кол-во продаж'].sum()
            stock_qty = category_df['Количество'].sum()
            store_qty_mix = category_df['Количество'].sum() / df.loc[df['Пол'] == "KIDS"]['Количество'].sum()
            sku_mix = category_df.loc[category_df['Актуальный заказ'] == "Актуально"]['SKU'].sum()/df.loc[(df['Актуальный заказ'] == "Актуально") & (df['Пол'] == "KIDS")]['SKU'].sum()
            act_sku = category_df.loc[category_df['Актуальный заказ'] == "Актуально"]['SKU'].sum()
            sku_a_mix = category_df.loc[(category_df['Grade'] == "A") & (category_df['Актуальный заказ'] == "Актуально")]['SKU'].sum() / act_sku
            sku_b_mix = category_df.loc[(category_df['Grade'] == "B") & (category_df['Актуальный заказ'] == "Актуально")]['SKU'].sum() / act_sku
            sku_c_mix = category_df.loc[(category_df['Grade'] == "C") & (category_df['Актуальный заказ'] == "Актуально")]['SKU'].sum() / act_sku
            sku_c = category_df.loc[(category_df['Grade'] == "C") & (category_df['Актуальный заказ'] == "Актуально")]['SKU'].sum()

            df_ttl_category_ttl = pd.DataFrame({
                "Category": [category],
                "NS": [ns],
                "NS MIX": [ns_mix],
                "NS RD MIX": [ns_mix],
                "NS QTY": [ns_qty],
                "NS QTY MIX": [ns_qty_mix],
                "STOCK QTY": [stock_qty],
                "STOCK QTY MIX": [store_qty_mix],
                "SKU": [act_sku],
                "SKU MIX": [sku_mix],
                "SKU A MIX": [sku_a_mix],
                "SKU B MIX": [sku_b_mix],
                "SKU C MIX": [sku_c_mix],
                "SKU C": [sku_c]
            })
            cat_df.append(df_ttl_category_ttl)

        self.store_by_kids_cat["Все магазины"] = cat_df

        for store in self.store_name:
            cat_store_df = []
            store_df = df.loc[(df['Магазин'] == store) & (df['Пол'] == "KIDS")]
            ns = store_df['Сумма со скидкой'].sum()
            ns_mix = 1
            ns_qty = store_df['Кол-во продаж'].sum()
            ns_qty_mix = 1
            stock_qty = store_df['Количество'].sum()
            store_qty_mix = 1
            sku_mix = 1
            act_sku = store_df.loc[store_df['Актуальный заказ'] == "Актуально"]['SKU'].sum()
            sku_a_mix = store_df.loc[(store_df['Grade'] == "A") & (store_df['Актуальный заказ'] == "Актуально")]['SKU'].sum() / act_sku
            sku_b_mix = store_df.loc[(store_df['Grade'] == "B") & (store_df['Актуальный заказ'] == "Актуально")]['SKU'].sum() / act_sku
            sku_c_mix = store_df.loc[(store_df['Grade'] == "C") & (store_df['Актуальный заказ'] == "Актуально")]['SKU'].sum() / act_sku
            sku_c = store_df.loc[(store_df['Grade'] == "C") & (store_df['Актуальный заказ'] == "Актуально")]['SKU'].sum()

            df_cat_ttl = pd.DataFrame({
                "Category": ["TTL"],
                "NS": [ns],
                "NS MIX": [ns_mix],
                "NS RD MIX": [ns_mix],
                "NS QTY": [ns_qty],
                "NS QTY MIX": [ns_qty_mix],
                "STOCK QTY": [stock_qty],
                "STOCK QTY MIX": [store_qty_mix],
                "SKU": [act_sku],
                "SKU MIX": [sku_mix],
                "SKU A MIX": [sku_a_mix],
                "SKU B MIX": [sku_b_mix],
                "SKU C MIX": [sku_c_mix],
                "SKU C": [sku_c]
            })

            for category in self.category_name:
                category_df = store_df.loc[(store_df['Категория'] == category) & (store_df['Пол'] == "KIDS")]
                category1_df = df.loc[(df['Категория'] == category) & (df['Пол'] == "KIDS")]
                ns_rd_mix = category1_df['Сумма со скидкой'].sum() / df.loc[df['Пол'] == "KIDS"]['Сумма со скидкой'].sum()
                ns = category_df['Сумма со скидкой'].sum()
                ns_mix = category_df['Сумма со скидкой'].sum() / store_df['Сумма со скидкой'].sum()
                ns_qty = category_df['Кол-во продаж'].sum()
                ns_qty_mix = category_df['Кол-во продаж'].sum() / store_df['Кол-во продаж'].sum()
                stock_qty = category_df['Количество'].sum()
                store_qty_mix = category_df['Количество'].sum() / store_df['Количество'].sum()
                sku_mix = category_df.loc[category_df['Актуальный заказ'] == "Актуально"]['SKU'].sum()/store_df.loc[store_df['Актуальный заказ'] == "Актуально"]['SKU'].sum()
                act_sku = category_df.loc[category_df['Актуальный заказ'] == "Актуально"]['SKU'].sum()
                sku_a_mix = category_df.loc[(category_df['Grade'] == "A") & (category_df['Актуальный заказ'] == "Актуально")]['SKU'].sum() / act_sku
                sku_b_mix = category_df.loc[(category_df['Grade'] == "B") & (category_df['Актуальный заказ'] == "Актуально")]['SKU'].sum() / act_sku
                sku_c_mix = category_df.loc[(category_df['Grade'] == "C") & (category_df['Актуальный заказ'] == "Актуально")]['SKU'].sum() / act_sku
                sku_c = category_df.loc[(category_df['Grade'] == "C") & (category_df['Актуальный заказ'] == "Актуально")]['SKU'].sum()

                df_gender_bystore = pd.DataFrame({
                    "Category": [category],
                    "NS": [ns],
                    "NS MIX": [ns_mix],
                    "NS RD MIX": [ns_rd_mix],
                    "NS QTY": [ns_qty],
                    "NS QTY MIX": [ns_qty_mix],
                    "STOCK QTY": [stock_qty],
                    "STOCK QTY MIX": [store_qty_mix],
                    "SKU": [act_sku],
                    "SKU MIX": [sku_mix],
                    "SKU A MIX": [sku_a_mix],
                    "SKU B MIX": [sku_b_mix],
                    "SKU C MIX": [sku_c_mix],
                    "SKU C": [sku_c]
                })
                cat_store_df.append(df_gender_bystore)
            cat_store_df.append(df_cat_ttl)
            self.store_by_kids_cat[store] = cat_store_df

    def get_by_model(self, df):
        numpy.seterr(divide='ignore', invalid='ignore')
        self.store_name = df['Магазин'].unique()
        self.category_name = df['Модель ANTA'].unique()

        cat_df = []
        for category in self.category_name:  # Тотал по всем магазинам
            category_df = df.loc[df['Модель ANTA'] == category]
            ns = category_df['Сумма со скидкой'].sum()
            ns_mix = category_df['Сумма со скидкой'].sum() / df['Сумма со скидкой'].sum()
            ns_qty = category_df['Кол-во продаж'].sum()
            ns_qty_mix = category_df['Кол-во продаж'].sum() / df['Кол-во продаж'].sum()
            stock_qty = category_df['Количество'].sum()
            store_qty_mix = category_df['Количество'].sum() / df['Количество'].sum()
            sku = category_df.loc[category_df['Актуальный заказ'] == "Актуально"]['SKU'].sum()
            sku_mix = category_df.loc[category_df['Актуальный заказ'] == "Актуально"]['SKU'].sum() / \
                      df.loc[df['Актуальный заказ'] == "Актуально"]['SKU'].sum()


            df_ttl_category_ttl = pd.DataFrame({
                "Category": [category],
                "NS": [ns],
                "NS MIX": [ns_mix],
                "NS RD MIX": [ns_mix],
                "NS QTY": [ns_qty],
                "NS QTY MIX": [ns_qty_mix],
                "STOCK QTY": [stock_qty],
                "STOCK QTY MIX": [store_qty_mix],
                "SKU": [sku],
                "SKU MIX": [sku_mix]
            })
            cat_df.append(df_ttl_category_ttl)

        self.store_by_models["Все магазины"] = cat_df

        for store in self.store_name:
            cat_store_df = []
            store_df = df.loc[df['Магазин'] == store]

            ns = store_df['Сумма со скидкой'].sum()
            ns_mix = 1
            ns_qty = store_df['Кол-во продаж'].sum()
            ns_qty_mix = 1
            stock_qty = store_df['Количество'].sum()
            store_qty_mix = 1
            sku = store_df.loc[store_df['Актуальный заказ'] == "Актуально"]['SKU'].sum()
            sku_mix = 1

            df_cat_ttl = pd.DataFrame({
                "Category": ["TTL"],
                "NS": [ns],
                "NS MIX": [ns_mix],
                "NS RD MIX": [ns_mix],
                "NS QTY": [ns_qty],
                "NS QTY MIX": [ns_qty_mix],
                "STOCK QTY": [stock_qty],
                "STOCK QTY MIX": [store_qty_mix],
                "SKU": [sku],
                "SKU MIX": [sku_mix]
            })

            for category in self.category_name:
                category_df = store_df.loc[store_df['Модель ANTA'] == category]
                category1_df = df.loc[df['Модель ANTA'] == category]
                ns_rd_mix = category1_df['Сумма со скидкой'].sum() / df['Сумма со скидкой'].sum()
                ns = category_df['Сумма со скидкой'].sum()
                ns_mix = category_df['Сумма со скидкой'].sum() / store_df['Сумма со скидкой'].sum()
                ns_qty = category_df['Кол-во продаж'].sum()
                ns_qty_mix = category_df['Кол-во продаж'].sum() / store_df['Кол-во продаж'].sum()
                stock_qty = category_df['Количество'].sum()
                store_qty_mix = category_df['Количество'].sum() / store_df['Количество'].sum()
                sku = category_df.loc[category_df['Актуальный заказ'] == "Актуально"]['SKU'].sum()
                sku_mix = category_df.loc[category_df['Актуальный заказ'] == "Актуально"]['SKU'].sum() / \
                          store_df.loc[store_df['Актуальный заказ'] == "Актуально"]['SKU'].sum()
                df_gender_bystore = pd.DataFrame({
                    "Category": [category],
                    "NS": [ns],
                    "NS MIX": [ns_mix],
                    "NS RD MIX": [ns_rd_mix],
                    "NS QTY": [ns_qty],
                    "NS QTY MIX": [ns_qty_mix],
                    "STOCK QTY": [stock_qty],
                    "STOCK QTY MIX": [store_qty_mix],
                    "SKU": [sku],
                    "SKU MIX": [sku_mix]
                })
                cat_store_df.append(df_gender_bystore)
            cat_store_df.append(df_cat_ttl)
            self.store_by_models[store] = cat_store_df

    def get_pr_type(self, df):
        numpy.seterr(divide='ignore', invalid='ignore')
        self.store_name = df['Магазин'].unique()
        self.category_name = df['Тип продукта'].unique()

        cat_df = []
        for category in self.category_name:  # Тотал по всем магазинам
            category_df = df.loc[df['Тип продукта'] == category]
            ns = category_df['Сумма со скидкой'].sum()
            ns_mix = category_df['Сумма со скидкой'].sum() / df['Сумма со скидкой'].sum()
            ns_qty = category_df['Кол-во продаж'].sum()
            ns_qty_mix = category_df['Кол-во продаж'].sum() / df['Кол-во продаж'].sum()
            stock_qty = category_df['Количество'].sum()
            store_qty_mix = category_df['Количество'].sum() / df['Количество'].sum()
            sku = category_df.loc[category_df['Актуальный заказ'] == "Актуально"]['SKU'].sum()
            sku_mix = category_df.loc[category_df['Актуальный заказ'] == "Актуально"]['SKU'].sum() / \
                      df.loc[df['Актуальный заказ'] == "Актуально"]['SKU'].sum()
            df_ttl_category_ttl = pd.DataFrame({
                "Category": [category],
                "NS": [ns],
                "NS MIX": [ns_mix],
                "NS RD MIX": [ns_mix],
                "NS QTY": [ns_qty],
                "NS QTY MIX": [ns_qty_mix],
                "STOCK QTY": [stock_qty],
                "STOCK QTY MIX": [store_qty_mix],
                "SKU": [sku],
                "SKU MIX": [sku_mix]
            })
            cat_df.append(df_ttl_category_ttl)

        self.store_by_product_type["Все магазины"] = cat_df

        for store in self.store_name:
            cat_store_df = []
            store_df = df.loc[df['Магазин'] == store]

            ns = store_df['Сумма со скидкой'].sum()
            ns_mix = 1
            ns_qty = store_df['Кол-во продаж'].sum()
            ns_qty_mix = 1
            stock_qty = store_df['Количество'].sum()
            store_qty_mix = 1
            sku = store_df.loc[store_df['Актуальный заказ'] == "Актуально"]['SKU'].sum()
            sku_mix = 1
            df_cat_ttl = pd.DataFrame({
                "Category": ["TTL"],
                "NS": [ns],
                "NS MIX": [ns_mix],
                "NS RD MIX": [ns_mix],
                "NS QTY": [ns_qty],
                "NS QTY MIX": [ns_qty_mix],
                "STOCK QTY": [stock_qty],
                "STOCK QTY MIX": [store_qty_mix],
                "SKU": [sku],
                "SKU MIX": [sku_mix]
            })

            for category in self.category_name:
                category_df = store_df.loc[store_df['Тип продукта'] == category]
                category1_df = df.loc[df['Тип продукта'] == category]
                ns_rd_mix = category1_df['Сумма со скидкой'].sum() / df['Сумма со скидкой'].sum()
                ns = category_df['Сумма со скидкой'].sum()
                ns_mix = category_df['Сумма со скидкой'].sum() / store_df['Сумма со скидкой'].sum()
                ns_qty = category_df['Кол-во продаж'].sum()
                ns_qty_mix = category_df['Кол-во продаж'].sum() / store_df['Кол-во продаж'].sum()
                stock_qty = category_df['Количество'].sum()
                store_qty_mix = category_df['Количество'].sum() / store_df['Количество'].sum()
                sku = category_df.loc[category_df['Актуальный заказ'] == "Актуально"]['SKU'].sum()
                sku_mix = category_df.loc[category_df['Актуальный заказ'] == "Актуально"]['SKU'].sum() / \
                          store_df.loc[store_df['Актуальный заказ'] == "Актуально"]['SKU'].sum()
                df_gender_bystore = pd.DataFrame({
                    "Category": [category],
                    "NS": [ns],
                    "NS MIX": [ns_mix],
                    "NS RD MIX": [ns_rd_mix],
                    "NS QTY": [ns_qty],
                    "NS QTY MIX": [ns_qty_mix],
                    "STOCK QTY": [stock_qty],
                    "STOCK QTY MIX": [store_qty_mix],
                    "SKU": [sku],
                    "SKU MIX": [sku_mix]
                })
                cat_store_df.append(df_gender_bystore)
            cat_store_df.append(df_cat_ttl)
            self.store_by_product_type[store] = cat_store_df

    def get_ftw_collect(self, df_1):
        numpy.seterr(divide='ignore', invalid='ignore')
        df = df_1.loc[df_1['Категория'] == "FTW"]
        self.store_name = df['Магазин'].unique()
        self.collections_name = df['Коллекция'].unique()

        cat_df = []
        for category in self.collections_name:  # Тотал по всем магазинам
            category_df = df.loc[(df['Коллекция'] == category) & (df['Категория'] == "FTW")]
            ns = category_df['Сумма со скидкой'].sum()
            ns_mix = category_df['Сумма со скидкой'].sum() / df.loc[df['Категория'] == "FTW"]['Сумма со скидкой'].sum()
            ns_qty = category_df['Кол-во продаж'].sum()
            ns_qty_mix = category_df['Кол-во продаж'].sum() / df.loc[df['Категория'] == "FTW"]['Кол-во продаж'].sum()
            stock_qty = category_df['Количество'].sum()
            store_qty_mix = category_df['Количество'].sum() / df.loc[df['Категория'] == "FTW"]['Количество'].sum()
            sku_mix = category_df.loc[category_df['Актуальный заказ'] == "Актуально"]['SKU'].sum() / \
                      df.loc[df['Актуальный заказ'] == "Актуально"]['SKU'].sum()
            act_sku = category_df.loc[category_df['Актуальный заказ'] == "Актуально"]['SKU'].sum()
            sku_a_mix = \
            category_df.loc[(category_df['Grade'] == "A") & (category_df['Актуальный заказ'] == "Актуально")][
                'SKU'].sum() / act_sku
            sku_b_mix = \
            category_df.loc[(category_df['Grade'] == "B") & (category_df['Актуальный заказ'] == "Актуально")][
                'SKU'].sum() / act_sku
            sku_c_mix = \
            category_df.loc[(category_df['Grade'] == "C") & (category_df['Актуальный заказ'] == "Актуально")][
                'SKU'].sum() / act_sku
            sku_c = category_df.loc[(category_df['Grade'] == "C") & (category_df['Актуальный заказ'] == "Актуально")][
                'SKU'].sum()

            df_ttl_category_ttl = pd.DataFrame({
                "Category": [category],
                "NS": [ns],
                "NS MIX": [ns_mix],
                "NS RD MIX": [ns_mix],
                "NS QTY": [ns_qty],
                "NS QTY MIX": [ns_qty_mix],
                "STOCK QTY": [stock_qty],
                "STOCK QTY MIX": [store_qty_mix],
                "SKU": [act_sku],
                "SKU MIX": [sku_mix],
                "SKU A MIX": [sku_a_mix],
                "SKU B MIX": [sku_b_mix],
                "SKU C MIX": [sku_c_mix],
                "SKU C": [sku_c]
            })
            cat_df.append(df_ttl_category_ttl)

        self.store_by_ftw_coll["Все магазины"] = cat_df

        for store in self.store_name:
            cat_store_df = []
            store_df = df.loc[(df['Магазин'] == store) & (df['Категория'] == "FTW")]

            ns = store_df['Сумма со скидкой'].sum()
            ns_mix = 1
            ns_qty = store_df['Кол-во продаж'].sum()
            ns_qty_mix = 1
            stock_qty = store_df['Количество'].sum()
            store_qty_mix = 1
            sku_mix = 1
            act_sku = store_df.loc[store_df['Актуальный заказ'] == "Актуально"]['SKU'].sum()
            sku_a_mix = store_df.loc[(store_df['Grade'] == "A") & (store_df['Актуальный заказ'] == "Актуально")][
                            'SKU'].sum() / act_sku
            sku_b_mix = store_df.loc[(store_df['Grade'] == "B") & (store_df['Актуальный заказ'] == "Актуально")][
                            'SKU'].sum() / act_sku
            sku_c_mix = store_df.loc[(store_df['Grade'] == "C") & (store_df['Актуальный заказ'] == "Актуально")][
                            'SKU'].sum() / act_sku
            sku_c = store_df.loc[(store_df['Grade'] == "C") & (store_df['Актуальный заказ'] == "Актуально")][
                'SKU'].sum()

            df_cat_ttl = pd.DataFrame({
                "Category": ["TTL"],
                "NS": [ns],
                "NS MIX": [ns_mix],
                "NS RD MIX": [ns_mix],
                "NS QTY": [ns_qty],
                "NS QTY MIX": [ns_qty_mix],
                "STOCK QTY": [stock_qty],
                "STOCK QTY MIX": [store_qty_mix],
                "SKU": [act_sku],
                "SKU MIX": [sku_mix],
                "SKU A MIX": [sku_a_mix],
                "SKU B MIX": [sku_b_mix],
                "SKU C MIX": [sku_c_mix],
                "SKU C": [sku_c]
            })
            self.collections_name = df.loc[(df['Магазин'] == store) & (df['Категория'] == "FTW")]['Коллекция'].unique()
            for category in self.collections_name:
                category_df = store_df.loc[(store_df['Коллекция'] == category) & (df['Категория'] == "FTW")]
                category1_df = df.loc[(df['Коллекция'] == category) & (df['Категория'] == "FTW")]
                ns_rd_mix = category1_df['Сумма со скидкой'].sum() / df.loc[df['Категория'] == "FTW"][
                    'Сумма со скидкой'].sum()
                ns = category_df['Сумма со скидкой'].sum()
                ns_mix = category_df['Сумма со скидкой'].sum() / store_df['Сумма со скидкой'].sum()
                ns_qty = category_df['Кол-во продаж'].sum()
                ns_qty_mix = category_df['Кол-во продаж'].sum() / store_df['Кол-во продаж'].sum()
                stock_qty = category_df['Количество'].sum()
                store_qty_mix = category_df['Количество'].sum() / store_df['Количество'].sum()
                sku_mix = category_df.loc[category_df['Актуальный заказ'] == "Актуально"]['SKU'].sum() / \
                          store_df.loc[store_df['Актуальный заказ'] == "Актуально"]['SKU'].sum()
                act_sku = category_df.loc[category_df['Актуальный заказ'] == "Актуально"]['SKU'].sum()
                sku_a_mix = \
                category_df.loc[(category_df['Grade'] == "A") & (category_df['Актуальный заказ'] == "Актуально")][
                    'SKU'].sum() / act_sku
                sku_b_mix = \
                category_df.loc[(category_df['Grade'] == "B") & (category_df['Актуальный заказ'] == "Актуально")][
                    'SKU'].sum() / act_sku
                sku_c_mix = \
                category_df.loc[(category_df['Grade'] == "C") & (category_df['Актуальный заказ'] == "Актуально")][
                    'SKU'].sum() / act_sku
                sku_c = \
                category_df.loc[(category_df['Grade'] == "C") & (category_df['Актуальный заказ'] == "Актуально")][
                    'SKU'].sum()

                df_gender_bystore = pd.DataFrame({
                    "Category": [category],
                    "NS": [ns],
                    "NS MIX": [ns_mix],
                    "NS RD MIX": [ns_rd_mix],
                    "NS QTY": [ns_qty],
                    "NS QTY MIX": [ns_qty_mix],
                    "STOCK QTY": [stock_qty],
                    "STOCK QTY MIX": [store_qty_mix],
                    "SKU": [act_sku],
                    "SKU MIX": [sku_mix],
                    "SKU A MIX": [sku_a_mix],
                    "SKU B MIX": [sku_b_mix],
                    "SKU C MIX": [sku_c_mix],
                    "SKU C": [sku_c]
                })
                cat_store_df.append(df_gender_bystore)
            cat_store_df.append(df_cat_ttl)
            self.store_by_ftw_coll[store] = cat_store_df

    def get_app_collect(self, df_1):
        numpy.seterr(divide='ignore', invalid='ignore')
        df = df_1.loc[df_1['Категория'] == "APP"]
        self.store_name = df['Магазин'].unique()
        self.collections_name = df['Коллекция'].unique()

        cat_df = []
        for category in self.collections_name:  # Тотал по всем магазинам
            category_df = df.loc[(df['Коллекция'] == category) & (df['Категория'] == "APP")]
            ns = category_df['Сумма со скидкой'].sum()
            ns_mix = category_df['Сумма со скидкой'].sum() / df.loc[df['Категория'] == "APP"]['Сумма со скидкой'].sum()
            ns_qty = category_df['Кол-во продаж'].sum()
            ns_qty_mix = category_df['Кол-во продаж'].sum() / df.loc[df['Категория'] == "APP"]['Кол-во продаж'].sum()
            stock_qty = category_df['Количество'].sum()
            store_qty_mix = category_df['Количество'].sum() / df.loc[df['Категория'] == "APP"]['Количество'].sum()
            sku_mix = category_df.loc[category_df['Актуальный заказ'] == "Актуально"]['SKU'].sum()/df.loc[df['Актуальный заказ'] == "Актуально"]['SKU'].sum()
            act_sku = category_df.loc[category_df['Актуальный заказ'] == "Актуально"]['SKU'].sum()
            sku_a_mix = category_df.loc[(category_df['Grade'] == "A") & (category_df['Актуальный заказ'] == "Актуально")]['SKU'].sum() / act_sku
            sku_b_mix = category_df.loc[(category_df['Grade'] == "B") & (category_df['Актуальный заказ'] == "Актуально")]['SKU'].sum() / act_sku
            sku_c_mix = category_df.loc[(category_df['Grade'] == "C") & (category_df['Актуальный заказ'] == "Актуально")]['SKU'].sum() / act_sku
            sku_c = category_df.loc[(category_df['Grade'] == "C") & (category_df['Актуальный заказ'] == "Актуально")]['SKU'].sum()

            df_ttl_category_ttl = pd.DataFrame({
                "Category": [category],
                "NS": [ns],
                "NS MIX": [ns_mix],
                "NS RD MIX": [ns_mix],
                "NS QTY": [ns_qty],
                "NS QTY MIX": [ns_qty_mix],
                "STOCK QTY": [stock_qty],
                "STOCK QTY MIX": [store_qty_mix],
                "SKU": [act_sku],
                "SKU MIX": [sku_mix],
                "SKU A MIX": [sku_a_mix],
                "SKU B MIX": [sku_b_mix],
                "SKU C MIX": [sku_c_mix],
                "SKU C": [sku_c]
            })
            cat_df.append(df_ttl_category_ttl)

        self.store_by_app_coll["Все магазины"] = cat_df

        for store in self.store_name:
            cat_store_df = []
            store_df = df.loc[(df['Магазин'] == store) & (df['Категория'] == "APP")]

            ns = store_df['Сумма со скидкой'].sum()
            ns_mix = 1
            ns_qty = store_df['Кол-во продаж'].sum()
            ns_qty_mix = 1
            stock_qty = store_df['Количество'].sum()
            store_qty_mix = 1
            sku_mix = 1
            act_sku = store_df.loc[store_df['Актуальный заказ'] == "Актуально"]['SKU'].sum()
            sku_a_mix = store_df.loc[(store_df['Grade'] == "A") & (store_df['Актуальный заказ'] == "Актуально")]['SKU'].sum() / act_sku
            sku_b_mix = store_df.loc[(store_df['Grade'] == "B") & (store_df['Актуальный заказ'] == "Актуально")]['SKU'].sum() / act_sku
            sku_c_mix = store_df.loc[(store_df['Grade'] == "C") & (store_df['Актуальный заказ'] == "Актуально")]['SKU'].sum() / act_sku
            sku_c = store_df.loc[(store_df['Grade'] == "C") & (store_df['Актуальный заказ'] == "Актуально")]['SKU'].sum()

            df_cat_ttl = pd.DataFrame({
                "Category": ["TTL"],
                "NS": [ns],
                "NS MIX": [ns_mix],
                "NS RD MIX": [ns_mix],
                "NS QTY": [ns_qty],
                "NS QTY MIX": [ns_qty_mix],
                "STOCK QTY": [stock_qty],
                "STOCK QTY MIX": [store_qty_mix],
                "SKU": [act_sku],
                "SKU MIX": [sku_mix],
                "SKU A MIX": [sku_a_mix],
                "SKU B MIX": [sku_b_mix],
                "SKU C MIX": [sku_c_mix],
                "SKU C": [sku_c]
            })
            self.collections_name = df.loc[(df['Магазин'] == store) & (df['Категория'] == "APP")]['Коллекция'].unique()
            for category in self.collections_name:
                category_df = store_df.loc[(store_df['Коллекция'] == category) & (df['Категория'] == "APP")]
                category1_df = df.loc[(df['Коллекция'] == category) & (df['Категория'] == "APP")]
                ns_rd_mix = category1_df['Сумма со скидкой'].sum() / df.loc[df['Категория'] == "APP"][
                    'Сумма со скидкой'].sum()
                ns = category_df['Сумма со скидкой'].sum()
                ns_mix = category_df['Сумма со скидкой'].sum() / store_df['Сумма со скидкой'].sum()
                ns_qty = category_df['Кол-во продаж'].sum()
                ns_qty_mix = category_df['Кол-во продаж'].sum() / store_df['Кол-во продаж'].sum()
                stock_qty = category_df['Количество'].sum()
                store_qty_mix = category_df['Количество'].sum() / store_df['Количество'].sum()
                sku_mix = category_df.loc[category_df['Актуальный заказ'] == "Актуально"]['SKU'].sum()/store_df.loc[store_df['Актуальный заказ'] == "Актуально"]['SKU'].sum()
                act_sku = category_df.loc[category_df['Актуальный заказ'] == "Актуально"]['SKU'].sum()
                sku_a_mix = category_df.loc[(category_df['Grade'] == "A") & (category_df['Актуальный заказ'] == "Актуально")]['SKU'].sum() / act_sku
                sku_b_mix = category_df.loc[(category_df['Grade'] == "B") & (category_df['Актуальный заказ'] == "Актуально")]['SKU'].sum() / act_sku
                sku_c_mix = category_df.loc[(category_df['Grade'] == "C") & (category_df['Актуальный заказ'] == "Актуально")]['SKU'].sum() / act_sku
                sku_c = category_df.loc[(category_df['Grade'] == "C") & (category_df['Актуальный заказ'] == "Актуально")]['SKU'].sum()

                df_gender_bystore = pd.DataFrame({
                    "Category": [category],
                    "NS": [ns],
                    "NS MIX": [ns_mix],
                    "NS RD MIX": [ns_rd_mix],
                    "NS QTY": [ns_qty],
                    "NS QTY MIX": [ns_qty_mix],
                    "STOCK QTY": [stock_qty],
                    "STOCK QTY MIX": [store_qty_mix],
                    "SKU": [act_sku],
                    "SKU MIX": [sku_mix],
                    "SKU A MIX": [sku_a_mix],
                    "SKU B MIX": [sku_b_mix],
                    "SKU C MIX": [sku_c_mix],
                    "SKU C": [sku_c]
                })
                cat_store_df.append(df_gender_bystore)
            cat_store_df.append(df_cat_ttl)
            self.store_by_app_coll[store] = cat_store_df

    def get_wmn_cat(self, df):
        numpy.seterr(divide='ignore', invalid='ignore')
        self.store_name = df['Магазин'].unique()
        self.category_name = df['Категория'].unique()

        cat_df = []
        for category in self.category_name:  # Тотал по всем магазинам
            category_df = df.loc[(df['Категория'] == category) & (df['Пол'] == "WOMEN")]
            ns = category_df['Сумма со скидкой'].sum()
            ns_mix = category_df['Сумма со скидкой'].sum() / df.loc[df['Пол'] == "WOMEN"]['Сумма со скидкой'].sum()
            ns_qty = category_df['Кол-во продаж'].sum()
            ns_qty_mix = category_df['Кол-во продаж'].sum() / df.loc[df['Пол'] == "WOMEN"]['Кол-во продаж'].sum()
            stock_qty = category_df['Количество'].sum()
            store_qty_mix = category_df['Количество'].sum() / df.loc[df['Пол'] == "WOMEN"]['Количество'].sum()
            sku_mix = category_df.loc[category_df['Актуальный заказ'] == "Актуально"]['SKU'].sum()/df.loc[(df['Актуальный заказ'] == "Актуально") & (df['Пол'] == "WOMEN")]['SKU'].sum()
            act_sku = category_df.loc[category_df['Актуальный заказ'] == "Актуально"]['SKU'].sum()
            sku_a_mix = category_df.loc[(category_df['Grade'] == "A") & (category_df['Актуальный заказ'] == "Актуально")]['SKU'].sum() / act_sku
            sku_b_mix = category_df.loc[(category_df['Grade'] == "B") & (category_df['Актуальный заказ'] == "Актуально")]['SKU'].sum() / act_sku
            sku_c_mix = category_df.loc[(category_df['Grade'] == "C") & (category_df['Актуальный заказ'] == "Актуально")]['SKU'].sum() / act_sku
            sku_c = category_df.loc[(category_df['Grade'] == "C") & (category_df['Актуальный заказ'] == "Актуально")]['SKU'].sum()

            df_ttl_category_ttl = pd.DataFrame({
                "Category": [category],
                "NS": [ns],
                "NS MIX": [ns_mix],
                "NS RD MIX": [ns_mix],
                "NS QTY": [ns_qty],
                "NS QTY MIX": [ns_qty_mix],
                "STOCK QTY": [stock_qty],
                "STOCK QTY MIX": [store_qty_mix],
                "SKU": [act_sku],
                "SKU MIX": [sku_mix],
                "SKU A MIX": [sku_a_mix],
                "SKU B MIX": [sku_b_mix],
                "SKU C MIX": [sku_c_mix],
                "SKU C": [sku_c]
            })
            cat_df.append(df_ttl_category_ttl)

        self.store_by_wmn_cat["Все магазины"] = cat_df

        for store in self.store_name:
            cat_store_df = []
            store_df = df.loc[(df['Магазин'] == store) & (df['Пол'] == "WOMEN")]

            ns = store_df['Сумма со скидкой'].sum()
            ns_mix = 1
            ns_qty = store_df['Кол-во продаж'].sum()
            ns_qty_mix = 1
            stock_qty = store_df['Количество'].sum()
            store_qty_mix = 1
            sku_mix = 1
            act_sku = store_df.loc[store_df['Актуальный заказ'] == "Актуально"]['SKU'].sum()
            sku_a_mix = store_df.loc[(store_df['Grade'] == "A") & (store_df['Актуальный заказ'] == "Актуально")]['SKU'].sum() / act_sku
            sku_b_mix = store_df.loc[(store_df['Grade'] == "B") & (store_df['Актуальный заказ'] == "Актуально")]['SKU'].sum() / act_sku
            sku_c_mix = store_df.loc[(store_df['Grade'] == "C") & (store_df['Актуальный заказ'] == "Актуально")]['SKU'].sum() / act_sku
            sku_c = store_df.loc[(store_df['Grade'] == "C") & (store_df['Актуальный заказ'] == "Актуально")]['SKU'].sum()

            df_cat_ttl = pd.DataFrame({
                "Category": ["TTL"],
                "NS": [ns],
                "NS MIX": [ns_mix],
                "NS RD MIX": [ns_mix],
                "NS QTY": [ns_qty],
                "NS QTY MIX": [ns_qty_mix],
                "STOCK QTY": [stock_qty],
                "STOCK QTY MIX": [store_qty_mix],
                "SKU": [act_sku],
                "SKU MIX": [sku_mix],
                "SKU A MIX": [sku_a_mix],
                "SKU B MIX": [sku_b_mix],
                "SKU C MIX": [sku_c_mix],
                "SKU C": [sku_c]
            })

            for category in self.category_name:
                numpy.seterr(divide='ignore')
                category_df = store_df.loc[(store_df['Категория'] == category) & (store_df['Пол'] == "WOMEN")]
                category1_df = df.loc[(df['Категория'] == category) & (df['Пол'] == "WOMEN")]
                ns_rd_mix = category1_df['Сумма со скидкой'].sum() / df.loc[df['Пол'] == "WOMEN"]['Сумма со скидкой'].sum()
                ns = category_df['Сумма со скидкой'].sum()
                ns_mix = category_df['Сумма со скидкой'].sum() / store_df['Сумма со скидкой'].sum()
                ns_qty = category_df['Кол-во продаж'].sum()
                ns_qty_mix = category_df['Кол-во продаж'].sum() / store_df['Кол-во продаж'].sum()
                stock_qty = category_df['Количество'].sum()
                store_qty_mix = category_df['Количество'].sum() / store_df['Количество'].sum()
                sku_mix = category_df.loc[category_df['Актуальный заказ'] == "Актуально"]['SKU'].sum()/store_df.loc[store_df['Актуальный заказ'] == "Актуально"]['SKU'].sum()
                act_sku = category_df.loc[category_df['Актуальный заказ'] == "Актуально"]['SKU'].sum()
                sku_a_mix = category_df.loc[(category_df['Grade'] == "A") & (category_df['Актуальный заказ'] == "Актуально")]['SKU'].sum() / act_sku
                sku_b_mix = category_df.loc[(category_df['Grade'] == "B") & (category_df['Актуальный заказ'] == "Актуально")]['SKU'].sum() / act_sku
                sku_c_mix = category_df.loc[(category_df['Grade'] == "C") & (category_df['Актуальный заказ'] == "Актуально")]['SKU'].sum() / act_sku
                sku_c = category_df.loc[(category_df['Grade'] == "C") & (category_df['Актуальный заказ'] == "Актуально")]['SKU'].sum()

                df_gender_bystore = pd.DataFrame({
                    "Category": [category],
                    "NS": [ns],
                    "NS MIX": [ns_mix],
                    "NS RD MIX": [ns_rd_mix],
                    "NS QTY": [ns_qty],
                    "NS QTY MIX": [ns_qty_mix],
                    "STOCK QTY": [stock_qty],
                    "STOCK QTY MIX": [store_qty_mix],
                    "SKU": [act_sku],
                    "SKU MIX": [sku_mix],
                    "SKU A MIX": [sku_a_mix],
                    "SKU B MIX": [sku_b_mix],
                    "SKU C MIX": [sku_c_mix],
                    "SKU C": [sku_c]
                })
                cat_store_df.append(df_gender_bystore)
            cat_store_df.append(df_cat_ttl)
            self.store_by_wmn_cat[store] = cat_store_df

    def get_men_cat(self, df):
        self.store_name = df['Магазин'].unique()
        self.category_name = df['Категория'].unique()

        cat_df = []
        for category in self.category_name:  # Тотал по всем магазинам
            category_df = df.loc[(df['Категория'] == category) & (df['Пол'] == "MEN")]
            ns = category_df['Сумма со скидкой'].sum()
            ns_mix = category_df['Сумма со скидкой'].sum() / df.loc[df['Пол'] == "MEN"]['Сумма со скидкой'].sum()
            ns_qty = category_df['Кол-во продаж'].sum()
            ns_qty_mix = category_df['Кол-во продаж'].sum() / df.loc[df['Пол'] == "MEN"]['Кол-во продаж'].sum()
            stock_qty = category_df['Количество'].sum()
            store_qty_mix = category_df['Количество'].sum() / df.loc[df['Пол'] == "MEN"]['Количество'].sum()
            sku_mix = category_df.loc[category_df['Актуальный заказ'] == "Актуально"]['SKU'].sum()/df.loc[(df['Актуальный заказ'] == "Актуально") & (df['Пол'] == "MEN")]['SKU'].sum()
            act_sku = category_df.loc[category_df['Актуальный заказ'] == "Актуально"]['SKU'].sum()
            sku_a_mix = category_df.loc[(category_df['Grade'] == "A") & (category_df['Актуальный заказ'] == "Актуально")]['SKU'].sum() / act_sku
            sku_b_mix = category_df.loc[(category_df['Grade'] == "B") & (category_df['Актуальный заказ'] == "Актуально")]['SKU'].sum() / act_sku
            sku_c_mix = category_df.loc[(category_df['Grade'] == "C") & (category_df['Актуальный заказ'] == "Актуально")]['SKU'].sum() / act_sku
            sku_c = category_df.loc[(category_df['Grade'] == "C") & (category_df['Актуальный заказ'] == "Актуально")]['SKU'].sum()

            df_ttl_category_ttl = pd.DataFrame({
                "Category": [category],
                "NS": [ns],
                "NS MIX": [ns_mix],
                "NS RD MIX": [ns_mix],
                "NS QTY": [ns_qty],
                "NS QTY MIX": [ns_qty_mix],
                "STOCK QTY": [stock_qty],
                "STOCK QTY MIX": [store_qty_mix],
                "SKU": [act_sku],
                "SKU MIX": [sku_mix],
                "SKU A MIX": [sku_a_mix],
                "SKU B MIX": [sku_b_mix],
                "SKU C MIX": [sku_c_mix],
                "SKU C": [sku_c]
            })
            cat_df.append(df_ttl_category_ttl)

        self.store_by_men_cat["Все магазины"] = cat_df

        for store in self.store_name:
            cat_store_df = []
            store_df = df.loc[(df['Магазин'] == store) & (df['Пол'] == "MEN")]
            ns = store_df['Сумма со скидкой'].sum()
            ns_mix = 1
            ns_qty = store_df['Кол-во продаж'].sum()
            ns_qty_mix = 1
            stock_qty = store_df['Количество'].sum()
            store_qty_mix = 1
            sku_mix = 1
            act_sku = store_df.loc[store_df['Актуальный заказ'] == "Актуально"]['SKU'].sum()
            sku_a_mix = store_df.loc[(store_df['Grade'] == "A") & (store_df['Актуальный заказ'] == "Актуально")]['SKU'].sum() / act_sku
            sku_b_mix = store_df.loc[(store_df['Grade'] == "B") & (store_df['Актуальный заказ'] == "Актуально")]['SKU'].sum() / act_sku
            sku_c_mix = store_df.loc[(store_df['Grade'] == "C") & (store_df['Актуальный заказ'] == "Актуально")]['SKU'].sum() / act_sku
            sku_c = store_df.loc[(store_df['Grade'] == "C") & (store_df['Актуальный заказ'] == "Актуально")]['SKU'].sum()

            df_cat_ttl = pd.DataFrame({
                "Category": ["TTL"],
                "NS": [ns],
                "NS MIX": [ns_mix],
                "NS RD MIX": [ns_mix],
                "NS QTY": [ns_qty],
                "NS QTY MIX": [ns_qty_mix],
                "STOCK QTY": [stock_qty],
                "STOCK QTY MIX": [store_qty_mix],
                "SKU": [act_sku],
                "SKU MIX": [sku_mix],
                "SKU A MIX": [sku_a_mix],
                "SKU B MIX": [sku_b_mix],
                "SKU C MIX": [sku_c_mix],
                "SKU C": [sku_c]
            })

            for category in self.category_name:
                category_df = store_df.loc[(store_df['Категория'] == category) & (store_df['Пол'] == "MEN")]
                category1_df = df.loc[(df['Категория'] == category) & (df['Пол'] == "MEN")]
                ns_rd_mix = category1_df['Сумма со скидкой'].sum() / df.loc[df['Пол'] == "MEN"]['Сумма со скидкой'].sum()
                ns = category_df['Сумма со скидкой'].sum()
                ns_mix = category_df['Сумма со скидкой'].sum() / store_df['Сумма со скидкой'].sum()
                ns_qty = category_df['Кол-во продаж'].sum()
                ns_qty_mix = category_df['Кол-во продаж'].sum() / store_df['Кол-во продаж'].sum()
                stock_qty = category_df['Количество'].sum()
                store_qty_mix = category_df['Количество'].sum() / store_df['Количество'].sum()
                sku_mix = category_df.loc[category_df['Актуальный заказ'] == "Актуально"]['SKU'].sum()/store_df.loc[store_df['Актуальный заказ'] == "Актуально"]['SKU'].sum()
                act_sku = category_df.loc[category_df['Актуальный заказ'] == "Актуально"]['SKU'].sum()
                sku_a_mix = category_df.loc[(category_df['Grade'] == "A") & (category_df['Актуальный заказ'] == "Актуально")]['SKU'].sum() / act_sku
                sku_b_mix = category_df.loc[(category_df['Grade'] == "B") & (category_df['Актуальный заказ'] == "Актуально")]['SKU'].sum() / act_sku
                sku_c_mix = category_df.loc[(category_df['Grade'] == "C") & (category_df['Актуальный заказ'] == "Актуально")]['SKU'].sum() / act_sku
                sku_c = category_df.loc[(category_df['Grade'] == "C") & (category_df['Актуальный заказ'] == "Актуально")]['SKU'].sum()

                df_gender_bystore = pd.DataFrame({
                    "Category": [category],
                    "NS": [ns],
                    "NS MIX": [ns_mix],
                    "NS RD MIX": [ns_rd_mix],
                    "NS QTY": [ns_qty],
                    "NS QTY MIX": [ns_qty_mix],
                    "STOCK QTY": [stock_qty],
                    "STOCK QTY MIX": [store_qty_mix],
                    "SKU": [act_sku],
                    "SKU MIX": [sku_mix],
                    "SKU A MIX": [sku_a_mix],
                    "SKU B MIX": [sku_b_mix],
                    "SKU C MIX": [sku_c_mix],
                    "SKU C": [sku_c]
                })
                cat_store_df.append(df_gender_bystore)
            cat_store_df.append(df_cat_ttl)
            self.store_by_men_cat[store] = cat_store_df

    def get_category_df(self, df):
        self.store_name = df['Магазин'].sort_values().unique()
        self.category_name = df['Категория'].unique()

        cat_df = []
        for category in self.category_name:  # Тотал по всем магазинам
            category_df = df.loc[df['Категория'] == category]
            ns = category_df['Сумма со скидкой'].sum()
            ns_mix = category_df['Сумма со скидкой'].sum() / df['Сумма со скидкой'].sum()
            ns_qty = category_df['Кол-во продаж'].sum()
            ns_qty_mix = category_df['Кол-во продаж'].sum() / df['Кол-во продаж'].sum()
            stock_qty = category_df['Количество'].sum()
            store_qty_mix = category_df['Количество'].sum() / df['Количество'].sum()
            sku_mix = category_df.loc[category_df['Актуальный заказ'] == "Актуально"]['SKU'].sum()/df.loc[df['Актуальный заказ'] == "Актуально"]['SKU'].sum()
            act_sku = category_df.loc[category_df['Актуальный заказ'] == "Актуально"]['SKU'].sum()
            sku_a_mix = category_df.loc[(category_df['Grade'] == "A") & (category_df['Актуальный заказ'] == "Актуально")]['SKU'].sum() / act_sku
            sku_b_mix = category_df.loc[(category_df['Grade'] == "B") & (category_df['Актуальный заказ'] == "Актуально")]['SKU'].sum() / act_sku
            sku_c_mix = category_df.loc[(category_df['Grade'] == "C") & (category_df['Актуальный заказ'] == "Актуально")]['SKU'].sum() / act_sku
            sku_c = category_df.loc[(category_df['Grade'] == "C") & (category_df['Актуальный заказ'] == "Актуально")]['SKU'].sum()

            df_ttl_category_ttl = pd.DataFrame({
                "Category": [category],
                "NS": [ns],
                "NS MIX": [ns_mix],
                "NS RD MIX": [ns_mix],
                "NS QTY": [ns_qty],
                "NS QTY MIX": [ns_qty_mix],
                "STOCK QTY": [stock_qty],
                "STOCK QTY MIX": [store_qty_mix],
                "SKU": [act_sku],
                "SKU MIX": [sku_mix],
                "SKU A MIX": [sku_a_mix],
                "SKU B MIX": [sku_b_mix],
                "SKU C MIX": [sku_c_mix],
                "SKU C": [sku_c]
            })
            cat_df.append(df_ttl_category_ttl)

        self.store_by_category["Все магазины"] = cat_df

        for store in self.store_name:
            cat_store_df = []
            store_df = df.loc[df['Магазин'] == store]

            ns = store_df['Сумма со скидкой'].sum()
            ns_mix = 1
            ns_qty = store_df['Кол-во продаж'].sum()
            ns_qty_mix = 1
            stock_qty = store_df['Количество'].sum()
            store_qty_mix = 1
            sku_mix = 1
            act_sku = store_df.loc[store_df['Актуальный заказ'] == "Актуально"]['SKU'].sum()
            sku_a_mix = store_df.loc[(store_df['Grade'] == "A") & (store_df['Актуальный заказ'] == "Актуально")]['SKU'].sum() / act_sku
            sku_b_mix = store_df.loc[(store_df['Grade'] == "B") & (store_df['Актуальный заказ'] == "Актуально")]['SKU'].sum() / act_sku
            sku_c_mix = store_df.loc[(store_df['Grade'] == "C") & (store_df['Актуальный заказ'] == "Актуально")]['SKU'].sum() / act_sku
            sku_c = store_df.loc[(store_df['Grade'] == "C") & (store_df['Актуальный заказ'] == "Актуально")]['SKU'].sum()

            df_cat_ttl = pd.DataFrame({
                "Category": ["TTL"],
                "NS": [ns],
                "NS MIX": [ns_mix],
                "NS RD MIX": [ns_mix],
                "NS QTY": [ns_qty],
                "NS QTY MIX": [ns_qty_mix],
                "STOCK QTY": [stock_qty],
                "STOCK QTY MIX": [store_qty_mix],
                "SKU": [act_sku],
                "SKU MIX": [sku_mix],
                "SKU A MIX": [sku_a_mix],
                "SKU B MIX": [sku_b_mix],
                "SKU C MIX": [sku_c_mix],
                "SKU C": [sku_c]
            })

            for category in self.category_name:
                category_df = store_df.loc[store_df['Категория'] == category]
                category1_df = df.loc[df['Категория'] == category]
                ns_rd_mix = category1_df['Сумма со скидкой'].sum() / df['Сумма со скидкой'].sum()
                ns = category_df['Сумма со скидкой'].sum()
                ns_mix = category_df['Сумма со скидкой'].sum() / store_df['Сумма со скидкой'].sum()
                ns_qty = category_df['Кол-во продаж'].sum()
                ns_qty_mix = category_df['Кол-во продаж'].sum() / store_df['Кол-во продаж'].sum()
                stock_qty = category_df['Количество'].sum()
                store_qty_mix = category_df['Количество'].sum() / store_df['Количество'].sum()
                sku_mix = category_df.loc[category_df['Актуальный заказ'] == "Актуально"]['SKU'].sum()/store_df.loc[store_df['Актуальный заказ'] == "Актуально"]['SKU'].sum()
                act_sku = category_df.loc[category_df['Актуальный заказ'] == "Актуально"]['SKU'].sum()
                sku_a_mix = category_df.loc[(category_df['Grade'] == "A") & (category_df['Актуальный заказ'] == "Актуально")]['SKU'].sum() / act_sku
                sku_b_mix = category_df.loc[(category_df['Grade'] == "B") & (category_df['Актуальный заказ'] == "Актуально")]['SKU'].sum() / act_sku
                sku_c_mix = category_df.loc[(category_df['Grade'] == "C") & (category_df['Актуальный заказ'] == "Актуально")]['SKU'].sum() / act_sku
                sku_c = category_df.loc[(category_df['Grade'] == "C") & (category_df['Актуальный заказ'] == "Актуально")]['SKU'].sum()

                df_gender_bystore = pd.DataFrame({
                    "Category": [category],
                    "NS": [ns],
                    "NS MIX": [ns_mix],
                    "NS RD MIX": [ns_rd_mix],
                    "NS QTY": [ns_qty],
                    "NS QTY MIX": [ns_qty_mix],
                    "STOCK QTY": [stock_qty],
                    "STOCK QTY MIX": [store_qty_mix],
                    "SKU": [act_sku],
                    "SKU MIX": [sku_mix],
                    "SKU A MIX": [sku_a_mix],
                    "SKU B MIX": [sku_b_mix],
                    "SKU C MIX": [sku_c_mix],
                    "SKU C": [sku_c]
                })
                cat_store_df.append(df_gender_bystore)
            cat_store_df.append(df_cat_ttl)
            self.store_by_category[store] = cat_store_df
        
    def get_gender_df(self, df):
        self.store_name = df['Магазин'].unique()
        self.gender_name = df['Пол'].unique()

        gend_df = []
        for gender in self.gender_name:     # Тотал по всем магазинам
            gender_df = df.loc[df['Пол'] == gender]
            ns = gender_df['Сумма со скидкой'].sum()
            ns_mix = gender_df['Сумма со скидкой'].sum() / df['Сумма со скидкой'].sum()
            ns_qty = gender_df['Кол-во продаж'].sum()
            ns_qty_mix = gender_df['Кол-во продаж'].sum() / df['Кол-во продаж'].sum()
            stock_qty = gender_df['Количество'].sum()
            store_qty_mix = gender_df['Количество'].sum() / df['Количество'].sum()
            sku_mix = gender_df.loc[gender_df['Актуальный заказ'] == "Актуально"]['SKU'].sum()/df.loc[df['Актуальный заказ'] == "Актуально"]['SKU'].sum()
            act_sku = gender_df.loc[gender_df['Актуальный заказ'] == "Актуально"]['SKU'].sum()
            sku_a_mix = gender_df.loc[(gender_df['Grade'] == "A") & (gender_df['Актуальный заказ'] == "Актуально")]['SKU'].sum() / act_sku
            sku_b_mix = gender_df.loc[(gender_df['Grade'] == "B") & (gender_df['Актуальный заказ'] == "Актуально")]['SKU'].sum() / act_sku
            sku_c_mix = gender_df.loc[(gender_df['Grade'] == "C") & (gender_df['Актуальный заказ'] == "Актуально")]['SKU'].sum() / act_sku
            sku_c = gender_df.loc[(gender_df['Grade'] == "C") & (gender_df['Актуальный заказ'] == "Актуально")]['SKU'].sum()

            df_ttl_gender_ttl = pd.DataFrame({
                "Gender": [gender],
                "NS": [ns],
                "NS MIX": [ns_mix],
                "NS RD MIX": [ns_mix],
                "NS QTY": [ns_qty],
                "NS QTY MIX": [ns_qty_mix],
                "STOCK QTY": [stock_qty],
                "STOCK QTY MIX": [store_qty_mix],
                "SKU": [act_sku],
                "SKU MIX": [sku_mix],
                "SKU A MIX": [sku_a_mix],
                "SKU B MIX": [sku_b_mix],
                "SKU C MIX": [sku_c_mix],
                "SKU C": [sku_c]
            })
            gend_df.append(df_ttl_gender_ttl)

        self.store_by_gender["Все магазины"] = gend_df

        for store in self.store_name:
            gend_store_df = []
            store_df = df.loc[df['Магазин'] == store]

            ns = store_df['Сумма со скидкой'].sum()
            ns_mix = 1
            ns_qty = store_df['Кол-во продаж'].sum()
            ns_qty_mix = 1
            stock_qty = store_df['Количество'].sum()
            store_qty_mix = 1
            sku_mix = 1
            act_sku = store_df.loc[store_df['Актуальный заказ'] == "Актуально"]['SKU'].sum()
            sku_a_mix = store_df.loc[(store_df['Grade'] == "A") & (store_df['Актуальный заказ'] == "Актуально")]['SKU'].sum() / act_sku
            sku_b_mix = store_df.loc[(store_df['Grade'] == "B") & (store_df['Актуальный заказ'] == "Актуально")]['SKU'].sum() / act_sku
            sku_c_mix = store_df.loc[(store_df['Grade'] == "C") & (store_df['Актуальный заказ'] == "Актуально")]['SKU'].sum() / act_sku
            sku_c = store_df.loc[(store_df['Grade'] == "C") & (store_df['Актуальный заказ'] == "Актуально")]['SKU'].sum()

            df_gender_ttl = pd.DataFrame({
                "Gender": ["TTL"],
                "NS": [ns],
                "NS MIX": [ns_mix],
                "NS RD MIX": [ns_mix],
                "NS QTY": [ns_qty],
                "NS QTY MIX": [ns_qty_mix],
                "STOCK QTY": [stock_qty],
                "STOCK QTY MIX": [store_qty_mix],
                "SKU": [act_sku],
                "SKU MIX": [sku_mix],
                "SKU A MIX": [sku_a_mix],
                "SKU B MIX": [sku_b_mix],
                "SKU C MIX": [sku_c_mix],
                "SKU C": [sku_c]
            })

            for gender in self.gender_name:
                gender_df = store_df.loc[store_df['Пол'] == gender]
                gender1_df = df.loc[df['Пол'] == gender]
                ns_rd_mix = gender1_df['Сумма со скидкой'].sum() / df['Сумма со скидкой'].sum()
                ns = gender_df['Сумма со скидкой'].sum()
                ns_mix = gender_df['Сумма со скидкой'].sum()/store_df['Сумма со скидкой'].sum()
                ns_qty = gender_df['Кол-во продаж'].sum()
                ns_qty_mix = gender_df['Кол-во продаж'].sum()/store_df['Кол-во продаж'].sum()
                stock_qty = gender_df['Количество'].sum()
                store_qty_mix = gender_df['Количество'].sum()/store_df['Количество'].sum()
                act_sku = gender_df.loc[gender_df['Актуальный заказ'] == "Актуально"]['SKU'].sum()
                sku_mix = gender_df.loc[gender_df['Актуальный заказ'] == "Актуально"]['SKU'].sum()/store_df.loc[store_df['Актуальный заказ'] == "Актуально"]['SKU'].sum()
                sku_a_mix = gender_df.loc[(gender_df['Grade'] == "A") & (gender_df['Актуальный заказ'] == "Актуально")]['SKU'].sum()/act_sku
                sku_b_mix = gender_df.loc[(gender_df['Grade'] == "B") & (gender_df['Актуальный заказ'] == "Актуально")]['SKU'].sum()/act_sku
                sku_c_mix = gender_df.loc[(gender_df['Grade'] == "C") & (gender_df['Актуальный заказ'] == "Актуально")]['SKU'].sum()/act_sku
                sku_c = gender_df.loc[(gender_df['Grade'] == "C") & (gender_df['Актуальный заказ'] == "Актуально")]['SKU'].sum()

                df_gender_bystore = pd.DataFrame({
                    "Gender": [gender],
                    "NS": [ns],
                    "NS MIX": [ns_mix],
                    "NS RD MIX": [ns_rd_mix],
                    "NS QTY": [ns_qty],
                    "NS QTY MIX": [ns_qty_mix],
                    "STOCK QTY": [stock_qty],
                    "STOCK QTY MIX": [store_qty_mix],
                    "SKU": [act_sku],
                    "SKU MIX": [sku_mix],
                    "SKU A MIX": [sku_a_mix],
                    "SKU B MIX": [sku_b_mix],
                    "SKU C MIX": [sku_c_mix],
                    "SKU C": [sku_c]
                })
                gend_store_df.append(df_gender_bystore)
            gend_store_df.append(df_gender_ttl)
            self.store_by_gender[store] = gend_store_df


if __name__ == "__main__":
    start = Processing()
    start.process("TestFile.xlsx")
